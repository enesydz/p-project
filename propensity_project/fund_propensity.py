"""Fund propensity modelling primitives for customer-month data."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd


PRODUCT_CLASSES = ("para_piyasasi", "nitelikli")
KEY_COLUMNS = ("musteri_id", "anchor_month", "product_class")
SOURCE_KEY_COLUMNS = ("musteri_id", "month", "product_class")
ACTIVITY_WIDE_COLUMNS = {"para_piyasasi": "ppf_aktif", "nitelikli": "nf_aktif"}


@dataclass(frozen=True)
class PropensityConfig:
    """Business and modelling controls used by the notebook."""

    x_grid: tuple[int, ...] = (1, 3, 6)
    y_grid: tuple[int, ...] = (1, 3)
    r_grid: tuple[float, ...] = (0.10, 0.25, 0.50)
    product_classes: tuple[str, ...] = PRODUCT_CLASSES
    min_eligible: int = 100
    min_positive: int = 20
    top_k: tuple[float, ...] = (0.01, 0.05, 0.10, 0.20)
    random_seed: int = 42
    campaign_capacity: int | None = None
    missing_threshold: float = 0.95
    correlation_threshold: float = 0.95
    correlation_sample_size: int | None = 50000
    max_categorical_levels: int = 100
    max_categorical_ratio: float = 0.50
    outlier_lower_quantile: float = 0.01
    outlier_upper_quantile: float = 0.99
    add_missing_indicators: bool = True
    segment_column: str = "segment"
    segment_values: tuple[Any, ...] | None = None
    test_months: int = 1
    oot_months: int = 1
    min_positive_train: int = 5
    min_positive_test: int = 1
    min_positive_oot: int = 1
    rare_event_rate_threshold: float = 0.01
    inflation_reference_month: Any | None = None
    inflation_adjust_all_continuous: bool = True
    input_table_file: str | None = None
    input_table_customer_column: str = "musteri_id"
    input_table_date_column: str = "month"
    input_table_product_column: str | None = "product_class"
    input_table_activity_columns: dict[str, str] = field(default_factory=lambda: dict(ACTIVITY_WIDE_COLUMNS))
    input_table_buy_column: str = "buy_amount"
    input_table_sell_column: str = "sell_amount"
    input_table_fund_value_column: str = "fund_value"
    input_table_feature_columns: tuple[str, ...] | None = None
    input_table_feature_aggregation: str = "first"
    activity_table_file: str | None = None
    activity_table_customer_column: str = "musteri_id"
    activity_table_date_column: str = "month"
    activity_ppf_flag_column: str = "ppf_aktif"
    activity_nf_flag_column: str = "nf_aktif"
    fund_table_file: str | None = None
    fund_table_customer_column: str = "musteri_id"
    fund_table_date_column: str = "month"
    fund_table_product_flag_column: str = "para_flg"
    fund_table_value_column: str = "tutar"
    transaction_table_file: str | None = None
    transaction_table_customer_column: str = "musteri_id"
    transaction_table_date_column: str = "month"
    transaction_table_product_flag_column: str = "ppf_flg"
    transaction_table_buy_column: str = "alim_tutari"
    transaction_table_sell_column: str = "satim_tutari"
    inflation_table_file: str | None = None
    inflation_date_column: str = "month"
    inflation_value_column: str = "inflation_rate"
    feature_windows: tuple[int, ...] = (1, 3, 6)
    feature_ratio_epsilon: float = 1e-6


@dataclass
class SourceBundle:
    """Normalized sources expected by the pipeline."""

    customers: pd.DataFrame
    activity: pd.DataFrame
    flows: pd.DataFrame
    monthly_features: pd.DataFrame = field(default_factory=pd.DataFrame)
    inflation: pd.DataFrame = field(default_factory=pd.DataFrame)


def _month_start(values: Iterable[Any]) -> pd.Series:
    return pd.to_datetime(pd.Series(values), errors="coerce").dt.to_period("M").dt.to_timestamp()


def _require_columns(frame: pd.DataFrame, required: Iterable[str], name: str) -> None:
    missing = sorted(set(required) - set(frame.columns))
    if missing:
        raise ValueError(f"{name} eksik kolonlar: {missing}")


def _column_token(value: Any) -> str:
    return str(value).strip().lower().replace("ı", "i").replace("ş", "s").replace("ğ", "g").replace("ü", "u").replace("ö", "o").replace("ç", "c")


def _normalize_segment_value(value: Any) -> str:
    if pd.isna(value):
        return "unknown"
    numeric = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(numeric) and float(numeric).is_integer():
        return str(int(numeric))
    return str(value).strip()


def _adapt_frame_columns(frame: pd.DataFrame, aliases: dict[str, tuple[str, ...]]) -> pd.DataFrame:
    """Rename common source aliases without overwriting existing canonical columns."""

    result = frame.copy()
    normalized = {_column_token(column): column for column in result.columns}
    rename: dict[str, str] = {}
    for canonical, candidates in aliases.items():
        if canonical in result.columns:
            continue
        for candidate in (canonical, *candidates):
            source = normalized.get(_column_token(candidate))
            if source is not None and source not in rename:
                rename[source] = canonical
                break
    return result.rename(columns=rename)


def adapt_source_bundle(bundle: SourceBundle) -> SourceBundle:
    """Adapt common English/Turkish source aliases to the canonical schema."""

    common = {
        "musteri_id": ("customer_id", "client_id", "customer", "client", "musteri", "musteri_no", "customer_no"),
        "month": ("date", "tarih", "ay", "month_date", "period"),
    }
    customers = _adapt_frame_columns(bundle.customers, {"musteri_id": common["musteri_id"], "segment": ("customer_segment", "customer_segment_id", "musteri_segment", "musteri_segment_id", "segment_id", "group")})
    activity = _adapt_frame_columns(bundle.activity, {
        **common,
        "ppf_aktif": ("ppf_active", "ppf_activity", "para_piyasasi_aktif", "money_market_active", "ppf_flag"),
        "nf_aktif": ("nf_active", "nitelikli_active", "nitelikli_aktif", "qualified_fund_active", "nf_flag"),
        "product_class": ("product", "urun", "fund_class", "product_group"),
        "active_flag": ("active", "aktif", "is_active", "activity_flag"),
    })
    flows = _adapt_frame_columns(bundle.flows, {
        **common,
        "product_class": ("product", "urun", "fund_class", "product_group"),
        "buy_amount": ("buy", "purchase_amount", "purchase", "alis", "alis_tutari", "buy_amt"),
        "sell_amount": ("sell", "sale_amount", "sale", "satis", "satis_tutari", "sell_amt"),
        "fund_value": ("aum", "balance", "fund_balance", "portfolio_value", "fon_bakiyesi", "fund_amount"),
    })
    monthly_features = _adapt_frame_columns(bundle.monthly_features, common) if not bundle.monthly_features.empty else bundle.monthly_features.copy()
    inflation = _adapt_frame_columns(bundle.inflation, {"month": common["month"], "inflation_rate": ("inflation", "cpi", "monthly_inflation", "enflasyon")}) if not bundle.inflation.empty else bundle.inflation.copy()
    return SourceBundle(customers, activity, flows, monthly_features, inflation)


def _prepare_keyed_frame(frame: pd.DataFrame, name: str) -> pd.DataFrame:
    result = frame.copy()
    _require_columns(result, SOURCE_KEY_COLUMNS, name)
    result["month"] = _month_start(result["month"])
    if result["month"].isna().any():
        raise ValueError(f"{name}.month icinde parse edilemeyen tarih var")
    result["musteri_id"] = result["musteri_id"].astype(str)
    result["product_class"] = result["product_class"].astype(str)
    duplicates = result.duplicated(list(SOURCE_KEY_COLUMNS), keep=False)
    if duplicates.any():
        sample = result.loc[duplicates, list(SOURCE_KEY_COLUMNS)].head(3).to_dict("records")
        raise ValueError(f"{name} anahtarinda duplicate var, ornek: {sample}")
    return result


def normalize_activity_table(activity: pd.DataFrame) -> pd.DataFrame:
    """Normalize monthly wide activity flags into one row per product class.

    The source activity table may contain either the canonical long schema
    (``product_class``, ``active_flag``) or the bank-style wide schema with
    ``ppf_aktif`` and ``nf_aktif`` columns.
    """

    _require_columns(activity, ["musteri_id", "month"], "activity")
    source = activity.copy()
    source["musteri_id"] = source["musteri_id"].astype(str)
    source["month"] = _month_start(source["month"])
    if source["month"].isna().any():
        raise ValueError("activity.month icinde parse edilemeyen tarih var")

    if {"product_class", "active_flag"}.issubset(source.columns):
        normalized = source[["musteri_id", "month", "product_class", "active_flag"]].copy()
    else:
        missing = sorted(set(ACTIVITY_WIDE_COLUMNS.values()) - set(source.columns))
        if missing:
            raise ValueError(
                "activity long formatta olmali veya ppf_aktif ve nf_aktif kolonlarini icermeli; "
                f"eksik: {missing}"
            )
        normalized = source[["musteri_id", "month", *ACTIVITY_WIDE_COLUMNS.values()]].melt(
            id_vars=["musteri_id", "month"],
            value_vars=list(ACTIVITY_WIDE_COLUMNS.values()),
            var_name="activity_column",
            value_name="active_flag",
        )
        reverse_mapping = {column: product_class for product_class, column in ACTIVITY_WIDE_COLUMNS.items()}
        normalized["product_class"] = normalized["activity_column"].map(reverse_mapping)
        normalized = normalized.drop(columns="activity_column")

    normalized["product_class"] = normalized["product_class"].astype(str)
    normalized["active_flag"] = pd.to_numeric(normalized["active_flag"], errors="coerce")
    if normalized["active_flag"].isna().any() or ~normalized["active_flag"].isin([0, 1]).all():
        raise ValueError("activity flag'leri yalnizca 0 veya 1 olmali")
    duplicates = normalized.duplicated(list(SOURCE_KEY_COLUMNS), keep=False)
    if duplicates.any():
        sample = normalized.loc[duplicates, list(SOURCE_KEY_COLUMNS)].head(3).to_dict("records")
        raise ValueError(f"activity anahtarinda duplicate var, ornek: {sample}")
    return normalized


def validate_sources(bundle: SourceBundle, config: PropensityConfig) -> dict[str, Any]:
    """Validate normalized sources and return a compact quality report."""

    bundle = adapt_source_bundle(bundle)
    _require_columns(bundle.customers, ["musteri_id"], "customers")
    _require_columns(bundle.activity, ["musteri_id", "month"], "activity")
    _require_columns(bundle.flows, [*SOURCE_KEY_COLUMNS, "buy_amount", "sell_amount", "fund_value"], "flows")
    customer_ids = bundle.customers["musteri_id"].astype(str)
    if customer_ids.duplicated().any():
        raise ValueError("customers.musteri_id tekil olmali")

    activity = normalize_activity_table(bundle.activity)
    flows = _prepare_keyed_frame(bundle.flows, "flows")
    for name, frame in (("activity", activity), ("flows", flows)):
        unknown_classes = sorted(set(frame["product_class"]) - set(config.product_classes))
        if unknown_classes:
            raise ValueError(f"{name} bilinmeyen product_class: {unknown_classes}")
        unknown_customers = sorted(set(frame["musteri_id"]) - set(customer_ids))
        if unknown_customers:
            raise ValueError(f"{name} master disi musteri iceriyor: {unknown_customers[:5]}")

    if pd.to_numeric(activity["active_flag"], errors="coerce").isna().any():
        raise ValueError("activity.active_flag numeric olmali")
    for column in ("buy_amount", "sell_amount", "fund_value"):
        values = pd.to_numeric(flows[column], errors="coerce")
        if values.isna().any() or (values < 0).any():
            raise ValueError(f"flows.{column} numeric ve negatif olmayan degerlerden olusmali")

    return {
        "customer_count": int(customer_ids.nunique()),
        "activity_rows": int(len(activity)),
        "flow_rows": int(len(flows)),
        "activity_month_min": activity["month"].min(),
        "activity_month_max": activity["month"].max(),
        "flow_month_min": flows["month"].min(),
        "flow_month_max": flows["month"].max(),
        "active_rate": float(pd.to_numeric(activity["active_flag"]).mean()),
    }


def build_canonical_panel(bundle: SourceBundle, config: PropensityConfig) -> pd.DataFrame:
    """Build a dense customer-month-product panel with zero-filled activity."""

    bundle = adapt_source_bundle(bundle)
    validate_sources(bundle, config)
    customers = bundle.customers.copy()
    customers["musteri_id"] = customers["musteri_id"].astype(str)
    if "segment" not in customers.columns:
        customers["segment"] = "unknown"
    customers[config.segment_column] = customers["segment"].map(_normalize_segment_value)
    if config.segment_column != "segment":
        customers = customers.drop(columns=["segment"])
    customers = customers.drop_duplicates("musteri_id")
    activity = normalize_activity_table(bundle.activity)
    flows = _prepare_keyed_frame(bundle.flows, "flows")
    activity["active_flag"] = pd.to_numeric(activity["active_flag"], errors="coerce").fillna(0)
    for column in ("buy_amount", "sell_amount", "fund_value"):
        flows[column] = pd.to_numeric(flows[column], errors="coerce").fillna(0.0)

    all_months = pd.concat([activity["month"], flows["month"]], ignore_index=True).dropna()
    if all_months.empty:
        raise ValueError("activity ve flows en az bir ay icermeli")
    months = pd.date_range(all_months.min(), all_months.max(), freq="MS")
    month_frame = pd.DataFrame({"month": months, "_join": 1})
    class_frame = pd.DataFrame({"product_class": list(config.product_classes), "_join": 1})
    customers["_join"] = 1
    panel = customers.merge(month_frame, on="_join").merge(class_frame, on="_join").drop(columns="_join")
    panel = panel.merge(activity[list(SOURCE_KEY_COLUMNS) + ["active_flag"]], on=list(SOURCE_KEY_COLUMNS), how="left", validate="one_to_one")
    panel = panel.merge(flows[list(SOURCE_KEY_COLUMNS) + ["buy_amount", "sell_amount", "fund_value"]], on=list(SOURCE_KEY_COLUMNS), how="left", validate="one_to_one")
    panel["active_flag"] = pd.to_numeric(panel["active_flag"], errors="coerce").fillna(0).astype("int8")
    for column in ("buy_amount", "sell_amount", "fund_value"):
        panel[column] = pd.to_numeric(panel[column], errors="coerce").fillna(0.0).astype("float64")
    panel["net_buy"] = panel["buy_amount"] - panel["sell_amount"]

    if not bundle.monthly_features.empty:
        monthly_features = bundle.monthly_features.copy()
        _require_columns(monthly_features, ["musteri_id", "month"], "monthly_features")
        monthly_features["musteri_id"] = monthly_features["musteri_id"].astype(str)
        monthly_features["month"] = _month_start(monthly_features["month"])
        if monthly_features.duplicated(["musteri_id", "month"]).any():
            raise ValueError("monthly_features anahtarinda duplicate var")
        panel = panel.merge(monthly_features, on=["musteri_id", "month"], how="left", validate="many_to_one")

    if not bundle.inflation.empty:
        inflation = bundle.inflation.copy()
        _require_columns(inflation, ["month", "inflation_rate"], "inflation")
        inflation["month"] = _month_start(inflation["month"])
        if inflation.duplicated(["month"]).any():
            raise ValueError("inflation ay anahtarinda duplicate var")
        inflation["inflation_rate"] = pd.to_numeric(inflation["inflation_rate"], errors="coerce")
        panel = panel.merge(inflation[["month", "inflation_rate"]], on="month", how="left", validate="many_to_one")
    else:
        panel["inflation_rate"] = np.nan
    return panel.sort_values(list(SOURCE_KEY_COLUMNS)).reset_index(drop=True)


def _future_sum(panel: pd.DataFrame, column: str, horizon: int) -> pd.Series:
    grouped = panel.groupby(["musteri_id", "product_class"], sort=False)[column]
    return pd.concat([grouped.shift(-offset) for offset in range(1, horizon + 1)], axis=1).sum(axis=1, min_count=horizon)


def _rolling_sum(panel: pd.DataFrame, column: str, window: int) -> pd.Series:
    return panel.groupby(["musteri_id", "product_class"], sort=False)[column].transform(lambda values: values.rolling(window, min_periods=window).sum())


def _rolling_max(panel: pd.DataFrame, column: str, window: int) -> pd.Series:
    return panel.groupby(["musteri_id", "product_class"], sort=False)[column].transform(lambda values: values.rolling(window, min_periods=window).max())


def build_target_table(panel: pd.DataFrame, config: PropensityConfig) -> pd.DataFrame:
    """Create auditable newsell and upsell labels for every parameter set."""

    _require_columns(panel, [*SOURCE_KEY_COLUMNS, "active_flag", "buy_amount", "sell_amount", "fund_value", "net_buy"], "panel")
    ordered = panel.sort_values(list(SOURCE_KEY_COLUMNS)).reset_index(drop=True).copy()
    max_month = ordered["month"].max()
    rows: list[pd.DataFrame] = []
    for product_class in config.product_classes:
        scoped = ordered[ordered["product_class"] == product_class].copy()
        for x_window in config.x_grid:
            recent_activity = _rolling_max(scoped, "active_flag", x_window)
            for y_window in config.y_grid:
                future_buy = _future_sum(scoped, "buy_amount", y_window)
                future_net = _future_sum(scoped, "net_buy", y_window)
                complete_future = scoped["month"] <= max_month - pd.DateOffset(months=y_window)
                newsell = scoped[["musteri_id", "month", "product_class"]].copy()
                newsell["model_type"] = "newsell"
                newsell["x_window"] = x_window
                newsell["y_window"] = y_window
                newsell["r_threshold"] = np.nan
                newsell["eligible"] = (recent_activity == 0) & complete_future
                newsell["target"] = np.where(newsell["eligible"], (future_buy.fillna(0) > 0).astype("int8"), np.nan)
                newsell["future_buy_amount"] = future_buy
                newsell["future_net_buy_amount"] = future_net
                newsell["anchor_fund_value"] = scoped["fund_value"].to_numpy()
                rows.append(newsell)
                for threshold in config.r_grid:
                    upsell = newsell.copy()
                    upsell["model_type"] = "upsell"
                    upsell["r_threshold"] = threshold
                    upsell["eligible"] = (scoped["active_flag"] == 1) & (scoped["fund_value"] > 0) & complete_future
                    uplift_ratio = future_net / scoped["fund_value"].replace(0, np.nan)
                    upsell["uplift_ratio"] = uplift_ratio
                    upsell["target"] = np.where(upsell["eligible"], (uplift_ratio >= threshold).fillna(False).astype("int8"), np.nan)
                    rows.append(upsell)
    result = pd.concat(rows, ignore_index=True).rename(columns={"month": "anchor_month"})
    result["anchor_month"] = pd.to_datetime(result["anchor_month"])
    result["eligible"] = result["eligible"].astype(bool)
    return result


def _inflation_reference_month(panel: pd.DataFrame, config: PropensityConfig) -> pd.Timestamp:
    if config.inflation_reference_month is not None:
        reference = pd.to_datetime(config.inflation_reference_month, errors="coerce")
        if pd.isna(reference):
            raise ValueError("inflation_reference_month parse edilemedi")
        return reference.to_period("M").to_timestamp()
    y_window = min(config.y_grid) if config.y_grid else config.oot_months
    return pd.to_datetime(panel["month"]).max() - pd.DateOffset(months=int(y_window))


def _build_inflation_frame(panel: pd.DataFrame, config: PropensityConfig) -> tuple[pd.DataFrame, pd.Timestamp]:
    reference_month = _inflation_reference_month(panel, config)
    months = pd.DataFrame({"month": sorted(pd.to_datetime(panel["month"]).dropna().unique())})
    if "inflation_rate" not in panel.columns:
        months["inflation_rate"] = 0.0
    else:
        rates = panel[["month", "inflation_rate"]].drop_duplicates("month").copy()
        rates["month"] = pd.to_datetime(rates["month"])
        rates["inflation_rate"] = pd.to_numeric(rates["inflation_rate"], errors="coerce").fillna(0.0)
        months = months.merge(rates, on="month", how="left", validate="one_to_one")
        months["inflation_rate"] = months["inflation_rate"].fillna(0.0)
    months = months.sort_values("month").reset_index(drop=True)
    months["inflation_index"] = (1.0 + months["inflation_rate"]).cumprod()
    reference_rows = months[months["month"] <= reference_month]
    if reference_rows.empty:
        reference_index = float(months["inflation_index"].iloc[0])
    else:
        reference_index = float(reference_rows["inflation_index"].iloc[-1])
    months["inflation_factor_to_oot"] = reference_index / months["inflation_index"].replace(0, np.nan)
    months["inflation_reference_month"] = reference_month
    return months, reference_month


def _continuous_inflation_columns(frame: pd.DataFrame, config: PropensityConfig) -> list[str]:
    excluded_tokens = ("flag", "active", "rate", "ratio", "pct", "percent", "count", "index", "factor", "month", "year")
    excluded_columns = {"musteri_id", "segment", "product_class"}
    columns: list[str] = []
    for column in frame.columns:
        if column in excluded_columns or not pd.api.types.is_numeric_dtype(frame[column]):
            continue
        token = str(column).lower()
        if any(excluded in token for excluded in excluded_tokens):
            continue
        unique_values = set(pd.to_numeric(frame[column], errors="coerce").dropna().unique().tolist())
        if unique_values and unique_values.issubset({0, 1}):
            continue
        if config.inflation_adjust_all_continuous:
            columns.append(column)
    return columns


def build_inflation_audit(panel: pd.DataFrame, features: pd.DataFrame, config: PropensityConfig) -> pd.DataFrame:
    """Report the inflation index and the continuous columns moved to the OOT date."""

    inflation_frame, reference_month = _build_inflation_frame(panel, config)
    adjusted_columns = _continuous_inflation_columns(panel, config)
    audit = inflation_frame.copy()
    audit["inflation_reference_month"] = reference_month
    audit["adjusted_continuous_columns"] = ", ".join(adjusted_columns)
    audit["adjusted_feature_count"] = len(adjusted_columns)
    audit["feature_table_row_count"] = len(features)
    return audit


def build_feature_table(panel: pd.DataFrame, config: PropensityConfig) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Create anchor-only features and a machine-readable lineage table."""

    ordered = panel.sort_values(list(SOURCE_KEY_COLUMNS)).reset_index(drop=True).copy()
    features = ordered.copy()
    inflation_frame, reference_month = _build_inflation_frame(features, config)
    features = features.merge(
        inflation_frame[["month", "inflation_rate", "inflation_factor_to_oot"]],
        on="month",
        how="left",
        validate="many_to_one",
        suffixes=("", "_source"),
    )
    if "inflation_rate_source" in features.columns:
        features["inflation_rate"] = features["inflation_rate_source"]
        features = features.drop(columns=["inflation_rate_source"])
    adjusted_columns = _continuous_inflation_columns(features, config)
    for column in adjusted_columns:
        features[column] = pd.to_numeric(features[column], errors="coerce") * features["inflation_factor_to_oot"]
    features["fund_value_log1p"] = np.log1p(features["fund_value"].clip(lower=0))
    features["net_buy_log1p_abs"] = np.sign(features["net_buy"]) * np.log1p(features["net_buy"].abs())
    grouped = features.groupby(["musteri_id", "product_class"], sort=False)
    features["fund_value_lag1"] = grouped["fund_value"].shift(1)
    features["fund_value_change_1m"] = features["fund_value"] - features["fund_value_lag1"]
    features["active_lag1"] = grouped["active_flag"].shift(1)
    lineage: list[dict[str, Any]] = [
        {"feature": "fund_value_log1p", "source": "fund_value", "lookback_months": 0, "as_of": "anchor_month", "transformation": "log1p_after_nonnegative_clip", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month"},
        {"feature": "net_buy_log1p_abs", "source": "net_buy", "lookback_months": 0, "as_of": "anchor_month", "transformation": "signed_log1p", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month"},
        {"feature": "fund_value_lag1", "source": "fund_value", "lookback_months": 1, "as_of": "anchor_month", "transformation": "lag_1_month", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month_minus_1"},
        {"feature": "fund_value_change_1m", "source": "fund_value", "lookback_months": 1, "as_of": "anchor_month", "transformation": "month_over_month_difference", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month_minus_1_to_anchor_month"},
        {"feature": "active_lag1", "source": "active_flag", "lookback_months": 1, "as_of": "anchor_month", "transformation": "lag_1_month", "outlier_analysis": "binary_no_outlier_transform", "period_rule": "anchor_month_minus_1"},
    ]
    base_columns = {"musteri_id", "month", "product_class", config.segment_column, "active_flag", "buy_amount", "sell_amount", "fund_value", "net_buy", "inflation_rate", "inflation_factor_to_oot"}
    numeric_monthly_features = [
        column for column in features.columns
        if column not in base_columns and pd.api.types.is_numeric_dtype(features[column])
    ]
    for column in numeric_monthly_features:
        lag_column = f"{column}_delta_1m"
        features[lag_column] = features[column] - grouped[column].shift(1)
        lineage.append({"feature": lag_column, "source": column, "lookback_months": 1, "as_of": "anchor_month", "transformation": "month_over_month_difference", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month_minus_1_to_anchor_month"})
        for window in sorted(set(config.feature_windows)):
            if window < 1:
                continue
            rolling_mean = grouped[column].transform(lambda values, window=window: values.rolling(window, min_periods=window).mean())
            ratio_column = f"{column}_ratio_{window}m"
            features[ratio_column] = features[column] / (rolling_mean.abs() + config.feature_ratio_epsilon)
            lineage.append({"feature": ratio_column, "source": column, "lookback_months": window, "as_of": "anchor_month", "transformation": f"current_to_rolling_mean_ratio_{window}m", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": f"anchor_month_minus_{window - 1}_to_anchor_month"})
        if 12 in config.feature_windows:
            seasonal_column = f"{column}_seasonal_delta_12m"
            features[seasonal_column] = features[column] - grouped[column].shift(12)
            lineage.append({"feature": seasonal_column, "source": column, "lookback_months": 12, "as_of": "anchor_month", "transformation": "year_over_year_difference", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month_minus_12"})
    for window in sorted(set(config.x_grid)):
        for column in ("buy_amount", "sell_amount", "net_buy"):
            output_column = f"{column}_sum_{window}m"
            features[output_column] = _rolling_sum(features, column, window)
            lineage.append({"feature": output_column, "source": column, "lookback_months": window, "as_of": "anchor_month", "transformation": f"rolling_sum_{window}m", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": f"anchor_month_minus_{window - 1}_to_anchor_month"})
        output_column = f"active_max_{window}m"
        features[output_column] = _rolling_max(features, "active_flag", window)
        lineage.append({"feature": output_column, "source": "active_flag", "lookback_months": window, "as_of": "anchor_month", "transformation": f"rolling_max_{window}m", "outlier_analysis": "binary_no_outlier_transform", "period_rule": f"anchor_month_minus_{window - 1}_to_anchor_month"})

    if "inflation_rate" in features.columns:
        features["fund_value_real"] = features["fund_value"]
        lineage.extend([
            {"feature": "inflation_rate", "source": "inflation", "lookback_months": 1, "as_of": "anchor_month", "transformation": "raw_monthly_rate", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month"},
            {"feature": "inflation_factor_to_oot", "source": "inflation", "lookback_months": 0, "as_of": "anchor_month", "transformation": "common_oot_reference_index_divided_by_anchor_index", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": f"reference_{reference_month:%Y-%m}"},
            {"feature": "fund_value_real", "source": "fund_value+inflation", "lookback_months": 0, "as_of": "anchor_month", "transformation": "inflation_adjusted_to_common_oot_date", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": f"reference_{reference_month:%Y-%m}"},
        ])
    excluded = set(KEY_COLUMNS) | {"month", "product_class", "musteri_id", "active_flag", "buy_amount", "sell_amount", "fund_value", "net_buy", "inflation_factor_to_oot"}
    feature_columns = [column for column in features.columns if column not in excluded and column != config.segment_column]
    lineage_names = {item["feature"] for item in lineage}
    for column in feature_columns:
        if column not in lineage_names:
            lineage.append({"feature": column, "source": "monthly_feature", "lookback_months": 0, "as_of": "anchor_month", "transformation": "raw_as_of_anchor", "outlier_analysis": "train_quantile_clip_at_modeling", "period_rule": "anchor_month"})
    feature_table = features[["musteri_id", "month", "product_class", config.segment_column, *feature_columns]].rename(columns={"month": "anchor_month"})
    feature_table = feature_table.replace([np.inf, -np.inf], np.nan)
    lineage_frame = pd.DataFrame(lineage).drop_duplicates("feature").sort_values("feature").reset_index(drop=True)
    return feature_table, lineage_frame


def build_feature_engineering_audit(features: pd.DataFrame, feature_lineage: pd.DataFrame, config: PropensityConfig) -> pd.DataFrame:
    """Describe feature transformations and segment-specific distribution checks."""

    segment_column = config.segment_column
    segments = sorted(features[segment_column].dropna().unique().tolist(), key=str) if segment_column in features.columns else ["unknown"]
    rows: list[dict[str, Any]] = []
    lineage_by_feature = feature_lineage.set_index("feature").to_dict("index")
    feature_columns = [column for column in features.columns if column not in KEY_COLUMNS and column != segment_column]
    for segment in segments:
        scoped = features[features[segment_column].eq(segment)] if segment_column in features.columns else features
        for feature in feature_columns:
            values = pd.to_numeric(scoped[feature], errors="coerce" )
            numeric = values.notna().sum() > 0
            if numeric:
                lower = float(values.quantile(config.outlier_lower_quantile))
                upper = float(values.quantile(config.outlier_upper_quantile))
                outlier_count = int(((values < lower) | (values > upper)).sum())
            else:
                lower = np.nan
                upper = np.nan
                outlier_count = 0
            lineage = lineage_by_feature.get(feature, {})
            rows.append({
                "segment": segment,
                "feature": feature,
                "source": lineage.get("source", "unknown"),
                "transformation": lineage.get("transformation", "raw_as_of_anchor"),
                "lookback_months": lineage.get("lookback_months", 0),
                "period_rule": lineage.get("period_rule", "anchor_month"),
                "as_of": lineage.get("as_of", "anchor_month"),
                "outlier_analysis": lineage.get("outlier_analysis", "train_quantile_clip_at_modeling"),
                "segment_row_count": len(scoped),
                "missing_pct": float(scoped[feature].isna().mean()),
                "unique_count": int(scoped[feature].nunique(dropna=True)),
                "analysis_lower_quantile": lower,
                "analysis_upper_quantile": upper,
                "analysis_outlier_count": outlier_count,
            })
    return pd.DataFrame(rows)


def target_audit(targets: pd.DataFrame) -> pd.DataFrame:
    """Summarize eligible volume and target rate by model configuration."""

    keys = ["model_type", "product_class", "x_window", "y_window", "r_threshold"]
    summary = targets.groupby(keys, dropna=False).agg(
        total_count=("target", "size"),
        eligible_count=("eligible", "sum"),
        positive_count=("target", "sum"),
    ).reset_index()
    eligible = targets[targets["eligible"]].copy()
    date_summary = eligible.groupby(keys, dropna=False).agg(
        anchor_month_min=("anchor_month", "min"),
        anchor_month_max=("anchor_month", "max"),
    ).reset_index()
    result = summary.merge(date_summary, on=keys, how="left", validate="one_to_one")
    result["target_rate"] = result["positive_count"] / result["eligible_count"].replace(0, np.nan)
    return result.sort_values(keys, na_position="first").reset_index(drop=True)


def variable_quality_report(frame: pd.DataFrame, stage: str, table_name: str = "frame") -> pd.DataFrame:
    """Profile every column for stage-level data quality reporting."""

    rows: list[dict[str, Any]] = []
    row_count = len(frame)
    for column in frame.columns:
        series = frame[column]
        numeric = pd.to_numeric(series, errors="coerce")
        numeric_count = int(numeric.notna().sum())
        non_null = int(series.notna().sum())
        unique_count = int(series.nunique(dropna=True))
        missing_pct = float(series.isna().mean()) if row_count else 0.0
        zero_pct = float((numeric == 0).mean()) if numeric_count else np.nan
        negative_pct = float((numeric < 0).mean()) if numeric_count else np.nan
        inf_count = int(np.isinf(numeric).sum()) if numeric_count else 0
        rows.append({
            "stage": stage,
            "table": table_name,
            "variable": column,
            "dtype": str(series.dtype),
            "row_count": row_count,
            "non_null_count": non_null,
            "missing_count": row_count - non_null,
            "missing_pct": missing_pct,
            "unique_count": unique_count,
            "constant_flag": unique_count <= 1,
            "numeric_flag": numeric_count == non_null and non_null > 0,
            "zero_pct": zero_pct,
            "negative_pct": negative_pct,
            "inf_count": inf_count,
            "min": float(numeric.min()) if numeric_count else np.nan,
            "max": float(numeric.max()) if numeric_count else np.nan,
            "mean": float(numeric.mean()) if numeric_count else np.nan,
            "std": float(numeric.std()) if numeric_count > 1 else 0.0,
            "quality_status": "OK" if missing_pct == 0 and inf_count == 0 and not (numeric_count and (numeric < 0).any()) else "CHECK",
        })
    return pd.DataFrame(rows)


def build_model_performance_summary(metrics: pd.DataFrame) -> pd.DataFrame:
    """Aggregate technical model metrics for notebook and Excel charts."""

    if metrics is None or metrics.empty:
        return pd.DataFrame(columns=["evaluation_stage", "split", "chart_label", "model_count", "pr_auc_mean", "roc_auc_mean", "brier_mean", "prevalence_mean", "pr_auc_lift_mean"])
    usable = metrics[metrics["split"].isin(["train", "test", "oot"])].copy()
    if usable.empty:
        return pd.DataFrame(columns=["evaluation_stage", "split", "chart_label", "model_count", "pr_auc_mean", "roc_auc_mean", "brier_mean", "prevalence_mean", "pr_auc_lift_mean"])
    group_columns = [column for column in ["segment", "evaluation_stage", "split"] if column in usable.columns]
    summary = usable.groupby(group_columns, dropna=False).agg(
        model_count=("model_key", "nunique"),
        pr_auc_mean=("pr_auc", "mean"),
        roc_auc_mean=("roc_auc", "mean"),
        brier_mean=("brier", "mean"),
        prevalence_mean=("prevalence", "mean"),
        positive_count=("positive_count", "sum"),
        sample_count=("sample_count", "sum"),
    ).reset_index()
    summary["pr_auc_lift_mean"] = summary["pr_auc_mean"] / summary["prevalence_mean"].replace(0, np.nan)
    summary["chart_label"] = ""
    if "segment" in summary:
        summary["chart_label"] = "S" + summary["segment"].astype(str) + " / "
    summary["chart_label"] = summary["chart_label"] + summary["evaluation_stage"].astype(str) + " / " + summary["split"].astype(str)
    columns = ["segment", "evaluation_stage", "split", "chart_label", "model_count", "pr_auc_mean", "roc_auc_mean", "brier_mean", "prevalence_mean", "pr_auc_lift_mean", "positive_count", "sample_count"]
    columns = [column for column in columns if column in summary.columns]
    return summary[columns].sort_values([column for column in ["segment", "evaluation_stage", "split"] if column in summary.columns]).reset_index(drop=True)


def build_model_status_summary(metrics: pd.DataFrame, target_quality: pd.DataFrame | None = None) -> pd.DataFrame:
    """Keep every segment/configuration visible, including gated or skipped models."""

    if metrics is None or metrics.empty:
        return pd.DataFrame(columns=["segment", "model_key", "model_type", "product_class", "split", "evaluation_stage", "status", "error"])
    status = metrics.copy()
    status["status"] = np.select(
        [status["split"].isin(["train", "test", "oot"]), status["split"].eq("skipped"), status["split"].eq("error")],
        ["evaluated", "skipped", "error"],
        default="unknown",
    )
    columns = ["segment", "model_key", "model_name", "model_type", "product_class", "x_window", "y_window", "r_threshold", "split", "evaluation_stage", "status", "error"]
    columns = [column for column in columns if column in status.columns]
    return status[columns].drop_duplicates().sort_values([column for column in ["segment", "model_key", "split"] if column in status.columns]).reset_index(drop=True)


def build_overfit_audit(metrics: pd.DataFrame) -> pd.DataFrame:
    """Compare train/test/OOT rare-event performance and flag overfit risk."""

    columns = [
        "model_key", "model_name", "evaluation_stage", "train_pr_auc", "test_pr_auc", "oot_pr_auc",
        "train_test_pr_auc_gap", "test_oot_pr_auc_gap", "test_pr_auc_lift", "oot_pr_auc_lift",
        "overfit_status", "rare_event_performance_status",
    ]
    if metrics is None or metrics.empty:
        return pd.DataFrame(columns=columns)
    usable = metrics[metrics["split"].isin(["train", "test", "oot"])].copy()
    if usable.empty:
        return pd.DataFrame(columns=columns)
    index = [column for column in ["model_key", "model_name", "evaluation_stage"] if column in usable.columns]
    pivot = usable.pivot_table(index=index, columns="split", values=["pr_auc", "prevalence"], aggfunc="first")
    result = pivot.reset_index()
    result.columns = [
        "_".join(str(part) for part in column if str(part) != "") if isinstance(column, tuple) else str(column)
        for column in result.columns
    ]
    for split in ("train", "test", "oot"):
        for metric in ("pr_auc", "prevalence"):
            source_column = f"{metric}_{split}"
            target_column = f"{split}_{metric}"
            if source_column in result:
                result[target_column] = result[source_column]
    for column in ["train_pr_auc", "test_pr_auc", "oot_pr_auc", "test_prevalence", "oot_prevalence"]:
        if column not in result:
            result[column] = np.nan
    result["train_test_pr_auc_gap"] = result["train_pr_auc"] - result["test_pr_auc"]
    result["test_oot_pr_auc_gap"] = result["test_pr_auc"] - result["oot_pr_auc"]
    result["test_pr_auc_lift"] = result["test_pr_auc"] / result["test_prevalence"].replace(0, np.nan)
    result["oot_pr_auc_lift"] = result["oot_pr_auc"] / result["oot_prevalence"].replace(0, np.nan)
    result["overfit_status"] = np.select(
        [
            result["train_test_pr_auc_gap"] > 0.20,
            result["test_oot_pr_auc_gap"] > 0.20,
        ],
        ["CHECK_train_test_gap", "CHECK_test_oot_gap"],
        default="OK",
    )
    result["rare_event_performance_status"] = np.where(
        (result["test_pr_auc_lift"] >= 1.0) & (result["oot_pr_auc_lift"] >= 1.0),
        "PR_AUC_above_prevalence",
        "CHECK_PR_AUC_vs_prevalence",
    )
    return result[columns].sort_values(["evaluation_stage", "model_key"]).reset_index(drop=True)


def build_runtime_summary(
    config: PropensityConfig,
    quality: dict[str, Any] | None,
    panel: pd.DataFrame,
    targets: pd.DataFrame,
    features: pd.DataFrame,
    metrics: pd.DataFrame | None,
    optuna_trials: pd.DataFrame | None,
    model_feature_audit: pd.DataFrame | None,
    overfit_audit: pd.DataFrame | None,
    elapsed_seconds: float | None = None,
) -> pd.DataFrame:
    """Create a first-sheet operational and model-control summary."""

    audit = model_feature_audit if model_feature_audit is not None else pd.DataFrame()
    metric_frame = metrics if metrics is not None else pd.DataFrame()
    trial_frame = optuna_trials if optuna_trials is not None else pd.DataFrame()
    rows = [
        {"section": "runtime", "metric": "elapsed_seconds", "value": elapsed_seconds, "unit": "seconds", "status": "INFO", "details": "Notebook run duration"},
        {"section": "runtime", "metric": "customer_count", "value": (quality or {}).get("customer_count", np.nan), "unit": "count", "status": "INFO", "details": "Validated source customers"},
        {"section": "runtime", "metric": "panel_rows", "value": len(panel), "unit": "rows", "status": "INFO", "details": "Dense customer-month-product panel"},
        {"section": "runtime", "metric": "target_rows", "value": len(targets), "unit": "rows", "status": "INFO", "details": "Auditable target table"},
        {"section": "runtime", "metric": "feature_rows", "value": len(features), "unit": "rows", "status": "INFO", "details": "Anchor feature table"},
        {"section": "model_control", "metric": "trained_model_count", "value": int(metric_frame.loc[metric_frame.get("split", pd.Series(dtype=str)).isin(["train", "test", "oot"]), "model_key"].nunique()) if not metric_frame.empty and "model_key" in metric_frame else 0, "unit": "models", "status": "INFO", "details": "Models with evaluation metrics"},
        {"section": "model_control", "metric": "overfit_check_status", "value": "CHECK" if overfit_audit is not None and (overfit_audit["overfit_status"] != "OK").any() else "OK", "unit": "status", "status": "CHECK" if overfit_audit is not None and (overfit_audit["overfit_status"] != "OK").any() else "OK", "details": "Train-test and test-OOT PR-AUC gaps; threshold 0.20"},
        {"section": "model_control", "metric": "rare_event_pr_auc_status", "value": "CHECK" if overfit_audit is not None and (overfit_audit["rare_event_performance_status"] != "PR_AUC_above_prevalence").any() else "OK", "unit": "status", "status": "CHECK" if overfit_audit is not None and (overfit_audit["rare_event_performance_status"] != "PR_AUC_above_prevalence").any() else "OK", "details": "Test and OOT PR-AUC must beat prevalence baseline"},
        {"section": "preprocessing", "metric": "outlier_clip_feature_count", "value": int((audit.get("outlier_method", pd.Series(dtype=str)).astype(str).str.startswith("quantile_clip")).sum()), "unit": "feature-audit rows", "status": "APPLIED" if not audit.empty and (audit.get("outlier_method", pd.Series(dtype=str)).astype(str).str.startswith("quantile_clip")).any() else "NOT_APPLIED", "details": f"Train quantiles {config.outlier_lower_quantile:.2f}-{config.outlier_upper_quantile:.2f}"},
        {"section": "preprocessing", "metric": "outlier_train_value_count", "value": int(pd.to_numeric(audit.get("train_outlier_count", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()), "unit": "values", "status": "INFO", "details": "Values clipped at train-fitted bounds"},
        {"section": "preprocessing", "metric": "eliminated_feature_count", "value": int((audit.get("action", pd.Series(dtype=str)) == "eliminated").sum()), "unit": "feature-audit rows", "status": "INFO", "details": "Missingness, constant, cardinality or correlation filters"},
        {"section": "preprocessing", "metric": "missing_imputation_feature_count", "value": int((audit.get("transformation", pd.Series(dtype=str)).astype(str).str.contains("median_impute", na=False)).sum()), "unit": "feature-audit rows", "status": "APPLIED" if not audit.empty and audit.get("transformation", pd.Series(dtype=str)).astype(str).str.contains("median_impute", na=False).any() else "NOT_APPLIED", "details": "Median imputation fitted on train only"},
        {"section": "preprocessing", "metric": "optuna_trial_count", "value": len(trial_frame), "unit": "trials", "status": "INFO", "details": "Completed/pruned/failed trial records"},
    ]
    return pd.DataFrame(rows)


def build_general_summary(
    stage_summary: pd.DataFrame,
    target_quality: pd.DataFrame,
    metrics: pd.DataFrame,
    oot_scores: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Build high-level volume and rare-event counts for a general chart."""

    ready_count = int(target_quality["model_ready"].sum()) if target_quality is not None and not target_quality.empty and "model_ready" in target_quality else 0
    rare_count = int(target_quality["rare_event_flag"].sum()) if target_quality is not None and not target_quality.empty and "rare_event_flag" in target_quality else 0
    trained_count = int(metrics.loc[metrics["split"].isin(["train", "test", "oot"]), "model_key"].nunique()) if metrics is not None and not metrics.empty else 0
    oot_row_count = int(len(oot_scores)) if oot_scores is not None else 0
    rows = [
        {"metric": "target_configurations", "value": int(len(target_quality)) if target_quality is not None else 0, "unit": "count"},
        {"metric": "model_ready_configurations", "value": ready_count, "unit": "count"},
        {"metric": "rare_event_configurations", "value": rare_count, "unit": "count"},
        {"metric": "trained_segment_models", "value": trained_count, "unit": "count"},
        {"metric": "oot_score_rows", "value": oot_row_count, "unit": "count"},
    ]
    if stage_summary is not None and not stage_summary.empty:
        rows.append({"metric": "stage_output_rows_total", "value": int(stage_summary["output_rows"].sum()), "unit": "count"})
    return pd.DataFrame(rows)


def create_performance_figures(performance_summary: pd.DataFrame, general_summary: pd.DataFrame) -> dict[str, Any]:
    """Create technical and general performance figures for notebook output."""

    import matplotlib.pyplot as plt

    technical, axes = plt.subplots(1, 2, figsize=(14, 5))
    if performance_summary is None or performance_summary.empty:
        axes[0].text(0.5, 0.5, "Model metriği bulunmuyor", ha="center", va="center")
        axes[1].text(0.5, 0.5, "Model metriği bulunmuyor", ha="center", va="center")
    else:
        chart = performance_summary.copy()
        labels = chart["chart_label"] if "chart_label" in chart else chart["evaluation_stage"].astype(str) + " / " + chart["split"].astype(str)
        axes[0].bar(labels, chart["pr_auc_mean"], label="PR-AUC", color="#245b73", alpha=0.9)
        axes[0].set_title("Rare-event ranking: PR-AUC")
        axes[0].set_ylabel("PR-AUC; prevalence baseline ile karşılaştır")
        axes[0].tick_params(axis="x", rotation=55)
        axes[0].legend()
        axes[1].bar(labels, chart["roc_auc_mean"], label="ROC-AUC", color="#c45b3c", alpha=0.85)
        axes[1].set_title("Ranking: ROC-AUC")
        axes[1].set_ylabel("ROC-AUC")
        axes[1].tick_params(axis="x", rotation=55)
        axes[1].legend()
    technical.tight_layout()

    general, axis = plt.subplots(figsize=(11, 5))
    if general_summary is None or general_summary.empty:
        axis.text(0.5, 0.5, "Genel özet bulunmuyor", ha="center", va="center")
    else:
        axis.bar(general_summary["metric"], general_summary["value"], color="#245b73")
        axis.set_title("Genel pipeline özeti")
        axis.set_ylabel("Adet")
        axis.tick_params(axis="x", rotation=35)
    general.tight_layout()
    return {"technical": technical, "general": general}


def save_performance_charts(figures: dict[str, Any], output_dir: str | Path) -> dict[str, Path]:
    """Persist chart images next to the Excel report."""

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    paths = {
        "technical": output_path / "propensity_performance_technical.png",
        "general": output_path / "propensity_performance_general.png",
    }
    for key, path in paths.items():
        figures[key].savefig(path, dpi=160, bbox_inches="tight")
    return paths


def build_pipeline_audit(
    bundle: SourceBundle,
    panel: pd.DataFrame,
    targets: pd.DataFrame,
    features: pd.DataFrame,
    feature_lineage: pd.DataFrame,
    config: PropensityConfig,
    selected_feature_columns: list[str] | None = None,
    metrics: pd.DataFrame | None = None,
    optuna_trials: pd.DataFrame | None = None,
    oot_scores: pd.DataFrame | None = None,
    campaign_scores: pd.DataFrame | None = None,
    model_feature_audit: pd.DataFrame | None = None,
    feature_engineering_audit: pd.DataFrame | None = None,
    inflation_audit: pd.DataFrame | None = None,
    performance_summary: pd.DataFrame | None = None,
    general_summary: pd.DataFrame | None = None,
    overfit_audit: pd.DataFrame | None = None,
    runtime_summary: pd.DataFrame | None = None,
    model_status_summary: pd.DataFrame | None = None,
) -> dict[str, pd.DataFrame]:
    """Build detailed, sheet-ready audit tables for every pipeline stage."""

    source_frames = {
        "customers": bundle.customers,
        "activity": bundle.activity,
        "flows": bundle.flows,
        "monthly_features": bundle.monthly_features,
        "inflation": bundle.inflation,
    }
    quality_frames = [
        variable_quality_report(frame, "data_load", name)
        for name, frame in source_frames.items()
        if not frame.empty
    ]
    quality_frames.extend([
        variable_quality_report(panel, "preprocess", "canonical_panel"),
        variable_quality_report(features, "feature_engineering", "feature_table"),
    ])
    variable_quality = pd.concat(quality_frames, ignore_index=True) if quality_frames else pd.DataFrame()

    source_variables = []
    for table_name, frame in source_frames.items():
        for column in frame.columns:
            source_variables.append({
                "stage": "data_load",
                "table": table_name,
                "variable": column,
                "action": "loaded",
                "reason": "source variable",
                "next_stage": "preprocess",
            })
    preprocess_variables = []
    for column in bundle.activity.columns:
        if column in {"ppf_aktif", "nf_aktif"}:
            action = "transformed"
            reason = "wide activity flag product_class + active_flag formatina acildi"
        else:
            action = "retained" if column in {"musteri_id", "month"} else "not_in_panel"
            reason = "canonical panel anahtari" if action == "retained" else "activity normalizasyonu sonrasi panelde kullanilmadi"
        preprocess_variables.append({"stage": "preprocess", "table": "activity", "variable": column, "action": action, "reason": reason, "next_stage": "canonical_panel"})

    selected_feature_columns = selected_feature_columns or [column for column in features.columns if column not in KEY_COLUMNS and column != config.segment_column]
    base_columns = {"musteri_id", "month", "product_class", "active_flag", "buy_amount", "sell_amount", "fund_value", "net_buy"}
    feature_variables = []
    for column in panel.columns:
        if column == config.segment_column:
            action = "excluded_from_model_features"
            reason = "segment partition key; her segment icin ayri model kuruldu"
        elif column in base_columns:
            action = "excluded_from_model_features"
            reason = "panel key veya target/transaction base kolonu; feature tablosunda dogrudan modele verilmedi"
        elif column in selected_feature_columns:
            action = "retained_as_model_feature"
            reason = "quality filtresinden gecti ve modele aktarildi"
        elif column in features.columns:
            action = "eliminated_from_model_features"
            reason = "quality filtresi veya model feature secimi disinda birakildi"
        else:
            action = "not_available"
            reason = "feature tablosunda bulunmuyor"
        feature_variables.append({"stage": "feature_engineering", "table": "canonical_panel", "variable": column, "action": action, "reason": reason, "next_stage": "modelling"})
    for row in feature_lineage.to_dict("records"):
        feature = row.get("feature")
        feature_variables.append({"stage": "feature_engineering", "table": "feature_lineage", "variable": feature, "action": "retained_as_model_feature" if feature in selected_feature_columns else "eliminated_from_model_features", "reason": row.get("source") if feature in selected_feature_columns else "quality filtresi veya model feature secimi disinda birakildi", "next_stage": "modelling"})

    feature_quality = variable_quality[variable_quality["table"] == "feature_table"].copy()
    feature_quality["quality_filter_action"] = np.where(
        feature_quality["variable"].isin(selected_feature_columns), "retained_for_model", "eliminated_from_model"
    )
    feature_quality["quality_filter_reason"] = np.where(
        feature_quality["variable"].isin(selected_feature_columns), "constant degil ve model feature listesine alindi", "constant, key veya secim disi kolon"
    )
    quality_eliminated = feature_quality[~feature_quality["variable"].isin(selected_feature_columns)][["stage", "table", "variable", "quality_filter_action", "quality_filter_reason"]]
    eliminated = pd.concat([
        pd.DataFrame(preprocess_variables),
        pd.DataFrame(feature_variables),
        quality_eliminated.rename(columns={"quality_filter_action": "action", "quality_filter_reason": "reason"}).assign(next_stage="modelling"),
    ], ignore_index=True)
    eliminated = eliminated[eliminated["action"].astype(str).str.contains("excluded|not_|eliminated|transformed", na=False)].drop_duplicates(["stage", "table", "variable", "action"])

    operation_log = pd.DataFrame([
        {"stage": "data_load", "step": "source_read", "action": "read_and_validate", "input_rows": sum(len(frame) for frame in source_frames.values()), "output_rows": sum(len(frame) for frame in source_frames.values()), "status": "completed", "details": "Kaynak tablolar okundu; kolon ve kaynak kalite profili uretildi."},
        {"stage": "preprocess", "step": "activity_normalize", "action": "wide_to_long", "input_rows": len(bundle.activity), "output_rows": len(panel), "status": "completed", "details": "ppf_aktif/nf_aktif kolonlari product_class ve active_flag alanlarina acildi."},
        {"stage": "preprocess", "step": "panel_build", "action": "dense_panel", "input_rows": len(bundle.activity), "output_rows": len(panel), "status": "completed", "details": "Musteri-ay-urun anahtari tekillestirildi ve eksik aktivite aylarina 0 yazildi."},
        {"stage": "feature_engineering", "step": "anchor_features", "action": "build_features", "input_rows": len(panel), "output_rows": len(features), "status": "completed", "details": "Sadece anchor ayi ve geriye donuk rolling pencereler kullanildi."},
        {"stage": "feature_engineering", "step": "feature_selection", "action": "retain_or_eliminate", "input_rows": len(features.columns), "output_rows": len(selected_feature_columns), "status": "completed", "details": "Key, constant ve secim disi kolonlar model girdisinden cikarildi."},
        {"stage": "feature_engineering", "step": "model_preprocessing", "action": "fit_on_train_only", "input_rows": len(features.columns), "output_rows": len(model_feature_audit) if model_feature_audit is not None else 0, "status": "completed", "details": "Encoding, outlier clipping, missing imputation ve correlation filtresi her model splitinin train verisinde fit edildi."},
        {"stage": "modelling", "step": "time_split_and_tuning", "action": "optuna_lightgbm", "input_rows": int(targets["eligible"].sum()), "output_rows": len(oot_scores) if oot_scores is not None else 0, "status": "completed", "details": "Kronolojik train/validation/OOT split ve pruned Optuna tuning uygulandi."},
        {"stage": "reporting", "step": "export", "action": "write_excel", "input_rows": len(oot_scores) if oot_scores is not None else 0, "output_rows": len(campaign_scores) if campaign_scores is not None else 0, "status": "completed", "details": "Audit, model ve kampanya tablolari workbook sheet'lerine yazildi."},
    ])

    stage_summary = pd.DataFrame([
        {"stage_order": 1, "stage": "data_load", "input_rows": sum(len(frame) for frame in source_frames.values()), "output_rows": sum(len(frame) for frame in source_frames.values()), "variable_count": len(variable_quality[variable_quality["stage"] == "data_load"]), "description": "Kaynak tablolar okundu ve source quality profili uretildi."},
        {"stage_order": 2, "stage": "preprocess", "input_rows": len(bundle.activity), "output_rows": len(panel), "variable_count": len(variable_quality[variable_quality["stage"] == "preprocess"]), "description": "Aktivite flag'leri normalize edildi, dense canonical panel olusturuldu."},
        {"stage_order": 3, "stage": "feature_engineering", "input_rows": len(panel), "output_rows": len(features), "variable_count": len(features.columns), "description": "Anchor-only feature'lar ve feature lineage uretildi."},
        {"stage_order": 4, "stage": "modelling", "input_rows": int(targets["eligible"].sum()), "output_rows": len(oot_scores) if oot_scores is not None else 0, "variable_count": len(features.columns), "description": "Optuna-LightGBM zaman bazli modelleme ve OOT scoring yapildi."},
        {"stage_order": 5, "stage": "reporting", "input_rows": len(oot_scores) if oot_scores is not None else 0, "output_rows": len(campaign_scores) if campaign_scores is not None else 0, "variable_count": 0, "description": "Metrik, trial, skor, campaign ve audit tablolarinin export asamasi."},
    ])

    target_with_segment = targets.merge(
        features[[*KEY_COLUMNS, config.segment_column]],
        on=list(KEY_COLUMNS),
        how="left",
        validate="many_to_one",
    )
    target_quality = target_with_segment.groupby([config.segment_column, "model_type", "product_class", "x_window", "y_window", "r_threshold"], dropna=False).agg(
        total_rows=("target", "size"), eligible_count=("eligible", "sum"), target_non_null=("target", "count"), positive_count=("target", "sum"),
    ).reset_index()
    target_quality["target_rate"] = target_quality["positive_count"] / target_quality["target_non_null"].replace(0, np.nan)
    target_quality["rare_event_flag"] = target_quality["target_rate"] <= config.rare_event_rate_threshold
    target_quality["minimum_positive_required"] = config.min_positive
    target_quality["model_ready"] = (target_quality["eligible_count"] >= config.min_eligible) & (target_quality["positive_count"] >= config.min_positive)
    target_quality["decision_reason"] = np.where(
        target_quality["model_ready"],
        np.where(target_quality["rare_event_flag"], "rare-event kosullari ve minimum veri kosulu saglandi", "minimum eligible ve positive kosullari saglandi"),
        "minimum veri kosulu saglanmadi",
    )
    performance_summary = performance_summary if performance_summary is not None else build_model_performance_summary(metrics if metrics is not None else pd.DataFrame())
    general_summary = general_summary if general_summary is not None else build_general_summary(stage_summary, target_quality, metrics if metrics is not None else pd.DataFrame(), oot_scores)
    overfit_audit = overfit_audit if overfit_audit is not None else build_overfit_audit(metrics if metrics is not None else pd.DataFrame())
    model_status_summary = model_status_summary if model_status_summary is not None else build_model_status_summary(metrics if metrics is not None else pd.DataFrame(), target_quality)

    tables = {
        "stage_summary": stage_summary,
        "operation_log": operation_log,
        "variable_quality": variable_quality,
        "eliminated_variables": eliminated,
        "preprocess_actions": pd.DataFrame(preprocess_variables),
        "feature_lineage": feature_lineage,
        "feature_quality": feature_quality,
        "feature_engineering_audit": feature_engineering_audit if feature_engineering_audit is not None else pd.DataFrame(),
        "model_feature_audit": model_feature_audit if model_feature_audit is not None else pd.DataFrame(),
        "inflation_audit": inflation_audit if inflation_audit is not None else pd.DataFrame(),
        "target_quality": target_quality,
        "target_audit": target_audit(targets),
        "model_metrics": metrics if metrics is not None else pd.DataFrame(),
        "optuna_trials": optuna_trials if optuna_trials is not None else pd.DataFrame(),
        "oot_scores": oot_scores if oot_scores is not None else pd.DataFrame(),
        "campaign_scores": campaign_scores if campaign_scores is not None else pd.DataFrame(),
        "performance_summary": performance_summary,
        "general_summary": general_summary,
        "overfit_audit": overfit_audit,
        "runtime_summary": runtime_summary if runtime_summary is not None else pd.DataFrame(),
        "model_status_summary": model_status_summary,
    }
    return tables


def write_pipeline_excel(audit_tables: dict[str, pd.DataFrame], output_path: str | Path) -> Path:
    """Write one worksheet per pipeline audit/report table."""

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sheet_names = {
        "runtime_summary": "00_Runtime_Summary",
        "stage_summary": "00_Stage_Summary",
        "operation_log": "01_Operation_Log",
        "variable_quality": "02_Variable_Quality",
        "eliminated_variables": "03_Eliminated_Vars",
        "preprocess_actions": "04_Preprocess_Actions",
        "feature_lineage": "05_Feature_Lineage",
        "feature_quality": "06_Feature_Quality",
        "feature_engineering_audit": "07_Feature_Engineering",
        "model_feature_audit": "08_Model_Feature_Audit",
        "target_quality": "09_Target_Quality",
        "target_audit": "10_Target_Audit",
        "model_metrics": "11_Model_Metrics",
        "optuna_trials": "12_Optuna_Trials",
        "oot_scores": "13_OOT_Scores",
        "campaign_scores": "14_Campaign_Scores",
        "inflation_audit": "15_Inflation_Audit",
        "performance_summary": "16_Performance_Summary",
        "general_summary": "17_General_Summary",
        "overfit_audit": "18_Overfit_Audit",
        "model_status_summary": "20_Model_Status_Summary",
    }
    sheet_names["stage_summary"] = "01_Stage_Summary"
    sheet_names["operation_log"] = "02_Operation_Log"
    sheet_names["variable_quality"] = "03_Variable_Quality"
    sheet_names["eliminated_variables"] = "04_Eliminated_Vars"
    sheet_names["preprocess_actions"] = "05_Preprocess_Actions"
    sheet_names["feature_lineage"] = "06_Feature_Lineage"
    sheet_names["feature_quality"] = "07_Feature_Quality"
    sheet_names["feature_engineering_audit"] = "08_Feature_Engineering"
    sheet_names["model_feature_audit"] = "09_Model_Feature_Audit"
    sheet_names["target_quality"] = "10_Target_Quality"
    sheet_names["target_audit"] = "11_Target_Audit"
    sheet_names["model_metrics"] = "12_Model_Metrics"
    sheet_names["optuna_trials"] = "13_Optuna_Trials"
    sheet_names["oot_scores"] = "14_OOT_Scores"
    sheet_names["campaign_scores"] = "15_Campaign_Scores"
    sheet_names["inflation_audit"] = "16_Inflation_Audit"
    sheet_names["performance_summary"] = "17_Performance_Summary"
    sheet_names["general_summary"] = "18_General_Summary"
    sheet_names["overfit_audit"] = "19_Overfit_Audit"
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for key, sheet_name in sheet_names.items():
            frame = audit_tables.get(key, pd.DataFrame())
            if frame.empty:
                frame = pd.DataFrame([{
                    "report_status": "NO_RECORDS_AFTER_GATE",
                    "record_count": 0,
                    "details": "Bu aşamada kayıt oluşmadı; model status ve target quality sheet'lerinde eleme nedeni bulunur.",
                }])
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.book[sheet_name]
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for column_cells in worksheet.columns:
                width = min(max(max(len(str(cell.value or "")) for cell in column_cells) + 2, 10), 42)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width
        _add_excel_summary_charts(writer.book)
    return output_path


def _add_excel_summary_charts(workbook: Any) -> None:
    """Attach technical and general charts to the summary worksheets."""

    from openpyxl.chart import BarChart, LineChart, Reference

    performance_sheet = workbook["17_Performance_Summary"]
    headers = {performance_sheet.cell(1, index).value: index for index in range(1, performance_sheet.max_column + 1)}
    if performance_sheet.max_row > 1 and {"chart_label", "pr_auc_mean", "roc_auc_mean"}.issubset(headers):
        chart = BarChart()
        chart.title = "Rare-event PR-AUC by split"
        chart.y_axis.title = "PR-AUC"
        chart.x_axis.title = "Stage / split"
        chart.height = 8
        chart.width = 22
        chart.style = 10
        chart.legend.position = "b"
        data = Reference(performance_sheet, min_col=headers["pr_auc_mean"], max_col=headers["pr_auc_mean"], min_row=1, max_row=performance_sheet.max_row)
        categories = Reference(performance_sheet, min_col=headers["chart_label"], max_col=headers["chart_label"], min_row=2, max_row=performance_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        performance_sheet.add_chart(chart, "J2")
        roc_chart = BarChart()
        roc_chart.title = "ROC-AUC by split"
        roc_chart.y_axis.title = "ROC-AUC"
        roc_chart.height = 8
        roc_chart.width = 22
        roc_chart.style = 10
        roc_chart.legend.position = "b"
        roc_data = Reference(performance_sheet, min_col=headers["roc_auc_mean"], max_col=headers["roc_auc_mean"], min_row=1, max_row=performance_sheet.max_row)
        roc_chart.add_data(roc_data, titles_from_data=True)
        roc_chart.set_categories(categories)
        performance_sheet.add_chart(roc_chart, "J20")

    general_sheet = workbook["18_General_Summary"]
    if general_sheet.max_row > 1 and general_sheet.max_column >= 2:
        chart = BarChart()
        chart.title = "General Pipeline Summary"
        chart.y_axis.title = "Count"
        chart.height = 9
        chart.width = 17
        data = Reference(general_sheet, min_col=2, max_col=2, min_row=1, max_row=general_sheet.max_row)
        categories = Reference(general_sheet, min_col=1, max_col=1, min_row=2, max_row=general_sheet.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        general_sheet.add_chart(chart, "E2")


def make_time_split(
    target_frame: pd.DataFrame,
    test_months: int = 1,
    oot_months: int = 1,
    min_positive_train: int = 1,
    min_positive_test: int = 1,
    min_positive_oot: int = 1,
) -> dict[str, Any]:
    """Return chronological train/test/OOT windows with rare-event safeguards."""

    if test_months < 1 or oot_months < 1:
        raise ValueError("test_months ve oot_months en az 1 olmali")
    months = sorted(pd.to_datetime(target_frame.loc[target_frame["eligible"], "anchor_month"].unique()))
    required_months = test_months + oot_months + 1
    if len(months) < required_months:
        raise ValueError(f"Zaman split icin en az {required_months} etiketli ay gerekli, bulunan: {len(months)}")
    train_months = months[: -(test_months + oot_months)]
    test_window = months[-(test_months + oot_months) : -oot_months]
    oot_window = months[-oot_months:]
    train_positive = int(target_frame[target_frame["anchor_month"].isin(train_months)]["target"].sum())
    test_positive = int(target_frame[target_frame["anchor_month"].isin(test_window)]["target"].sum())
    oot_positive = int(target_frame[target_frame["anchor_month"].isin(oot_window)]["target"].sum())
    if train_positive < min_positive_train or test_positive < min_positive_test or oot_positive < min_positive_oot:
        raise ValueError(
            "split pozitif esigi saglanmadi: "
            f"train={train_positive} (min {min_positive_train}), "
            f"test={test_positive} (min {min_positive_test}), "
            f"oot={oot_positive} (min {min_positive_oot})"
        )
    return {
        "train_months": train_months,
        "test_months": test_window,
        "oot_months": oot_window,
        "train_positive": train_positive,
        "test_positive": test_positive,
        "oot_positive": oot_positive,
    }


def _safe_metric(function: Any, *args: Any, **kwargs: Any) -> float:
    try:
        return float(function(*args, **kwargs))
    except ValueError:
        return float("nan")


def evaluate_predictions(y_true: pd.Series, probability: np.ndarray, top_k: Iterable[float] = (0.01, 0.05, 0.10, 0.20)) -> dict[str, float]:
    """Evaluate rare-event ranking and probability quality."""

    from sklearn.metrics import average_precision_score, brier_score_loss, roc_auc_score
    y_array = np.asarray(y_true, dtype=int)
    probability = np.asarray(probability, dtype=float)
    order = np.argsort(-probability)
    prevalence = float(y_array.mean()) if len(y_array) else np.nan
    metrics = {
        "sample_count": int(len(y_array)),
        "positive_count": int(y_array.sum()),
        "prevalence": prevalence,
        "pr_auc": _safe_metric(average_precision_score, y_array, probability),
        "roc_auc": _safe_metric(roc_auc_score, y_array, probability),
        "brier": _safe_metric(brier_score_loss, y_array, probability),
    }
    for fraction in top_k:
        count = max(1, int(np.ceil(len(y_array) * fraction)))
        selected = y_array[order[:count]]
        precision = float(selected.mean()) if len(selected) else np.nan
        suffix = f"{fraction * 100:g}pct".replace(".", "_")
        metrics[f"precision_at_{suffix}"] = precision
        metrics[f"recall_at_{suffix}"] = float(selected.sum() / max(y_array.sum(), 1))
        metrics[f"lift_at_{suffix}"] = precision / prevalence if prevalence > 0 else np.nan
    return metrics


def fit_logistic_baseline(train_frame: pd.DataFrame, score_frame: pd.DataFrame, feature_columns: list[str], target_column: str = "target", random_seed: int = 42) -> tuple[Any, dict[str, float], pd.DataFrame]:
    """Fit a weighted logistic baseline."""

    from sklearn.compose import ColumnTransformer
    from sklearn.impute import SimpleImputer
    from sklearn.linear_model import LogisticRegression
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import OneHotEncoder, StandardScaler
    numeric_columns = [column for column in feature_columns if pd.api.types.is_numeric_dtype(train_frame[column])]
    categorical_columns = [column for column in feature_columns if column not in numeric_columns]
    transformers: list[tuple[str, Any, list[str]]] = []
    if numeric_columns:
        transformers.append(("numeric", Pipeline([( "imputer", SimpleImputer(strategy="median", add_indicator=True)), ("scaler", StandardScaler())]), numeric_columns))
    if categorical_columns:
        transformers.append(("categorical", Pipeline([( "imputer", SimpleImputer(strategy="most_frequent")), ("onehot", OneHotEncoder(handle_unknown="ignore", sparse_output=False))]), categorical_columns))
    preprocessor = ColumnTransformer(transformers=transformers, remainder="drop")
    model = Pipeline([( "preprocessor", preprocessor), ("classifier", LogisticRegression(class_weight="balanced", max_iter=500, random_state=random_seed))])
    model.fit(train_frame[feature_columns], train_frame[target_column].astype(int))
    probability = model.predict_proba(score_frame[feature_columns])[:, 1]
    metrics = evaluate_predictions(score_frame[target_column], probability)
    predictions = score_frame[["musteri_id", "anchor_month", "product_class"]].copy()
    predictions["probability"] = probability
    return model, metrics, predictions


def fit_hist_gradient_boosting(train_frame: pd.DataFrame, score_frame: pd.DataFrame, feature_columns: list[str], target_column: str = "target", random_seed: int = 42) -> tuple[Any, dict[str, float], pd.DataFrame]:
    """Fit a nonlinear scikit-learn challenger without requiring LightGBM."""

    from sklearn.compose import ColumnTransformer
    from sklearn.ensemble import HistGradientBoostingClassifier
    from sklearn.impute import SimpleImputer
    from sklearn.pipeline import Pipeline
    from sklearn.preprocessing import OneHotEncoder
    from sklearn.utils.class_weight import compute_sample_weight
    numeric_columns = [column for column in feature_columns if pd.api.types.is_numeric_dtype(train_frame[column])]
    categorical_columns = [column for column in feature_columns if column not in numeric_columns]
    transformers: list[tuple[str, Any, list[str]]] = []
    if numeric_columns:
        transformers.append(("numeric", Pipeline([( "imputer", SimpleImputer(strategy="median", add_indicator=True))]), numeric_columns))
    if categorical_columns:
        transformers.append(("categorical", Pipeline([( "imputer", SimpleImputer(strategy="most_frequent")), ("onehot", OneHotEncoder(handle_unknown="ignore", sparse_output=False))]), categorical_columns))
    preprocessor = ColumnTransformer(transformers=transformers, remainder="drop", sparse_threshold=0)
    model = Pipeline([( "preprocessor", preprocessor), ("classifier", HistGradientBoostingClassifier(max_iter=150, learning_rate=0.05, max_leaf_nodes=31, l2_regularization=1.0, random_state=random_seed))])
    sample_weight = compute_sample_weight("balanced", train_frame[target_column].astype(int))
    model.fit(train_frame[feature_columns], train_frame[target_column].astype(int), classifier__sample_weight=sample_weight)
    probability = model.predict_proba(score_frame[feature_columns])[:, 1]
    metrics = evaluate_predictions(score_frame[target_column], probability)
    predictions = score_frame[["musteri_id", "anchor_month", "product_class"]].copy()
    predictions["probability"] = probability
    return model, metrics, predictions


def _configuration_mask(targets: pd.DataFrame, model_type: str, product_class: str, x_window: int, y_window: int, r_threshold: float | None) -> pd.Series:
    mask = (
        targets["model_type"].eq(model_type)
        & targets["product_class"].eq(product_class)
        & targets["x_window"].eq(x_window)
        & targets["y_window"].eq(y_window)
    )
    if r_threshold is None:
        return mask & targets["r_threshold"].isna()
    return mask & targets["r_threshold"].eq(r_threshold)


def run_model_grid(
    targets: pd.DataFrame,
    features: pd.DataFrame,
    config: PropensityConfig,
    max_configurations: int | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any]]:
    """Run baseline and nonlinear models on chronological OOT splits.

    Configurations failing the minimum eligible/positive thresholds or lacking
    two classes in train/OOT are returned in the audit and skipped for training.
    """

    feature_columns = [column for column in features.columns if column not in KEY_COLUMNS and column != config.segment_column]
    audit = target_audit(targets)
    ready = audit[(audit["eligible_count"] >= config.min_eligible) & (audit["positive_count"] >= config.min_positive)].copy()
    if max_configurations is not None:
        ready = ready.head(max_configurations)
    metric_rows: list[dict[str, Any]] = []
    prediction_rows: list[pd.DataFrame] = []
    model_registry: dict[str, Any] = {}

    for row in ready.itertuples(index=False):
        threshold = None if pd.isna(row.r_threshold) else float(row.r_threshold)
        mask = _configuration_mask(targets, row.model_type, row.product_class, int(row.x_window), int(row.y_window), threshold)
        selected = targets.loc[mask & targets["eligible"]].copy()
        selected = selected.merge(features, on=list(KEY_COLUMNS), how="left", validate="one_to_one")
        selected = selected.dropna(subset=["target"])
        try:
            split = make_time_split(selected)
        except ValueError:
            continue
        train = selected[selected["anchor_month"].isin(split["train_months"])].copy()
        test = selected[selected["anchor_month"].isin(split["test_months"])].copy()
        oot = selected[selected["anchor_month"].isin(split["oot_months"])].copy()
        if train["target"].nunique() < 2 or test["target"].nunique() < 2 or oot["target"].nunique() < 2:
            continue
        model_key = f"{row.model_type}_{row.product_class}_x{row.x_window}_y{row.y_window}_r{threshold}"
        for model_name, fitter in (("logistic", fit_logistic_baseline), ("hist_gradient_boosting", fit_hist_gradient_boosting)):
            try:
                model, test_metrics, test_predictions = fitter(train, test, feature_columns, random_seed=config.random_seed)
                _, oot_metrics, oot_predictions = fitter(train, oot, feature_columns, random_seed=config.random_seed)
            except (ValueError, TypeError) as error:
                metric_rows.append({"model_key": model_key, "model_name": model_name, "split": "error", "error": str(error), "model_type": row.model_type, "product_class": row.product_class, "x_window": row.x_window, "y_window": row.y_window, "r_threshold": threshold})
                continue
            for split_name, metrics in (("test", test_metrics), ("oot", oot_metrics)):
                metric_rows.append({"model_key": model_key, "model_name": model_name, "split": split_name, "model_type": row.model_type, "product_class": row.product_class, "x_window": row.x_window, "y_window": row.y_window, "r_threshold": threshold, **metrics})
            oot_predictions["model_key"] = model_key
            oot_predictions["model_name"] = model_name
            oot_predictions["model_type"] = row.model_type
            oot_predictions["x_window"] = row.x_window
            oot_predictions["y_window"] = row.y_window
            oot_predictions["r_threshold"] = threshold
            prediction_rows.append(oot_predictions)
            model_registry[f"{model_key}_{model_name}"] = model

    metrics_frame = pd.DataFrame(metric_rows)
    predictions_frame = pd.concat(prediction_rows, ignore_index=True) if prediction_rows else pd.DataFrame()
    return metrics_frame, predictions_frame, model_registry


def fit_feature_contract(
    train_frame: pd.DataFrame,
    feature_columns: list[str],
    config: PropensityConfig,
) -> tuple[dict[str, Any], pd.DataFrame]:
    """Fit leakage-safe missing, outlier, correlation and encoding rules on train only."""

    report: list[dict[str, Any]] = []
    numeric_candidates: list[str] = []
    categorical_candidates: list[str] = []
    row_count = max(len(train_frame), 1)
    for column in feature_columns:
        if column not in train_frame.columns:
            report.append({"feature": column, "action": "eliminated", "reason": "train frame icinde bulunmuyor"})
            continue
        series = train_frame[column]
        missing_count = int(series.isna().sum())
        missing_pct = missing_count / row_count
        unique_count = int(series.nunique(dropna=True))
        base = {"feature": column, "dtype": str(series.dtype), "missing_pct": missing_pct, "unique_count": unique_count, "missing_count": missing_count, "outlier_lower": np.nan, "outlier_upper": np.nan, "train_outlier_count": 0, "encoding": "none", "transformation": "none", "outlier_method": "not_applicable", "correlation_method": "not_applicable", "fit_scope": "train_only"}
        if missing_pct > config.missing_threshold:
            report.append({**base, "action": "eliminated", "reason": f"missing_pct>{config.missing_threshold:.2f}"})
            continue
        if unique_count <= 1:
            report.append({**base, "action": "eliminated", "reason": "constant_or_empty"})
            continue
        if pd.api.types.is_numeric_dtype(series):
            numeric_candidates.append(column)
        else:
            if unique_count > config.max_categorical_levels or unique_count / row_count > config.max_categorical_ratio:
                report.append({**base, "action": "eliminated", "reason": "categorical_cardinality_too_high"})
                continue
            categorical_candidates.append(column)

    numeric_frame = train_frame[numeric_candidates].replace([np.inf, -np.inf], np.nan).copy()
    numeric_frame = numeric_frame.fillna(numeric_frame.median(numeric_only=True))
    correlation_drops: set[str] = set()
    correlation_scope = "all_train_rows" if config.correlation_sample_size is None or len(train_frame) <= config.correlation_sample_size else f"train_sample_{config.correlation_sample_size}"
    if len(numeric_candidates) > 1 and config.correlation_threshold < 1.0:
        correlation_frame = numeric_frame
        if config.correlation_sample_size is not None and len(correlation_frame) > config.correlation_sample_size:
            correlation_frame = correlation_frame.sample(config.correlation_sample_size, random_state=config.random_seed)
        correlation = correlation_frame.corr().abs()
        upper = correlation.where(np.triu(np.ones(correlation.shape), k=1).astype(bool))
        correlation_drops = {column for column in upper.columns if (upper[column] > config.correlation_threshold).any()}
    selected_numeric = [column for column in numeric_candidates if column not in correlation_drops]
    selected = set(selected_numeric) | set(categorical_candidates)
    bounds: dict[str, tuple[float, float]] = {}
    medians: dict[str, float] = {}
    categories: dict[str, list[str]] = {}
    for column in selected_numeric:
        values = pd.to_numeric(train_frame[column], errors="coerce").replace([np.inf, -np.inf], np.nan)
        median = float(values.median()) if values.notna().any() else 0.0
        lower = float(values.quantile(config.outlier_lower_quantile)) if values.notna().any() else median
        upper = float(values.quantile(config.outlier_upper_quantile)) if values.notna().any() else median
        if lower > upper:
            lower, upper = upper, lower
        bounds[column] = (lower, upper)
        medians[column] = median
    for column in categorical_candidates:
        categories[column] = sorted(train_frame[column].dropna().astype(str).unique().tolist())

    for item in report:
        if item["feature"] in correlation_drops:
            item.update({"action": "eliminated", "reason": f"high_correlation>{config.correlation_threshold:.2f}"})
    report_by_feature = {item["feature"]: item for item in report}
    for column in selected_numeric:
        lower, upper = bounds[column]
        values = pd.to_numeric(train_frame[column], errors="coerce")
        outlier_count = int(((values < lower) | (values > upper)).sum())
        report_by_feature[column] = {**report_by_feature.get(column, {"feature": column, "dtype": str(train_frame[column].dtype), "missing_pct": train_frame[column].isna().mean(), "unique_count": train_frame[column].nunique(dropna=True), "missing_count": train_frame[column].isna().sum()}), "outlier_lower": lower, "outlier_upper": upper, "train_outlier_count": outlier_count, "action": "retained", "reason": "numeric_clip_and_median_imputation", "encoding": "numeric", "transformation": "quantile_clip_then_median_impute", "outlier_method": f"quantile_clip_{config.outlier_lower_quantile:.2f}_{config.outlier_upper_quantile:.2f}", "correlation_method": f"drop_upper_triangle_over_{config.correlation_threshold:.2f};{correlation_scope}", "fit_scope": "train_only"}
    for column in categorical_candidates:
        report_by_feature[column] = {**report_by_feature.get(column, {"feature": column, "dtype": str(train_frame[column].dtype), "missing_pct": train_frame[column].isna().mean(), "unique_count": train_frame[column].nunique(dropna=True), "missing_count": train_frame[column].isna().sum()}), "action": "retained", "reason": "train_categories_only", "encoding": "one_hot", "transformation": "train_categories_only_one_hot", "outlier_method": "not_applicable", "correlation_method": "numeric_only_correlation_filter", "fit_scope": "train_only"}
    for column in correlation_drops:
        report_by_feature[column] = {**report_by_feature.get(column, {"feature": column}), "action": "eliminated", "reason": f"high_correlation>{config.correlation_threshold:.2f}", "encoding": "none", "transformation": "eliminated", "outlier_method": "not_applicable", "correlation_method": f"drop_upper_triangle_over_{config.correlation_threshold:.2f}", "fit_scope": "train_only"}
    contract = {"numeric": selected_numeric, "categorical": categorical_candidates, "bounds": bounds, "medians": medians, "categories": categories, "missing_indicators": [column for column in selected if config.add_missing_indicators and train_frame[column].isna().any()], "correlation_scope": correlation_scope}
    return contract, pd.DataFrame(list(report_by_feature.values())).sort_values("feature").reset_index(drop=True)


def transform_feature_contract(frame: pd.DataFrame, contract: dict[str, Any]) -> pd.DataFrame:
    """Apply a train-fitted feature contract to any later split or new dataset."""

    result = pd.DataFrame(index=frame.index)
    for column in contract["numeric"]:
        values = pd.to_numeric(frame[column], errors="coerce").replace([np.inf, -np.inf], np.nan)
        if column in contract["missing_indicators"]:
            result[f"{column}__missing"] = values.isna().astype("int8")
        lower, upper = contract["bounds"][column]
        result[column] = values.clip(lower, upper).fillna(contract["medians"][column]).astype("float32")
    for column in contract["categorical"]:
        values = frame[column].astype("string").fillna("__MISSING__")
        allowed = contract["categories"][column]
        for category in allowed:
            result[f"{column}__{category}"] = (values == category).astype("int8")
        if column in contract["missing_indicators"]:
            result[f"{column}__missing"] = frame[column].isna().astype("int8")
    return result.reset_index(drop=True)


def _encode_lgbm_frames(frames: list[pd.DataFrame], feature_columns: list[str], config: PropensityConfig) -> tuple[list[pd.DataFrame], list[str], pd.DataFrame]:
    """Fit encoding and robust feature rules on train, then transform all splits."""

    contract, audit = fit_feature_contract(frames[0], feature_columns, config)
    encoded = [transform_feature_contract(frame, contract) for frame in frames]
    return encoded, list(encoded[0].columns), audit


def fit_optuna_lightgbm(
    train_frame: pd.DataFrame,
    validation_frame: pd.DataFrame,
    score_frame: pd.DataFrame,
    feature_columns: list[str],
    n_trials: int = 30,
    timeout: int | None = 600,
    random_seed: int = 42,
    config: PropensityConfig | None = None,
) -> tuple[dict[str, Any], dict[str, float], pd.DataFrame, pd.DataFrame]:
    """Tune LightGBM with Optuna's MedianPruner and iteration-level pruning."""

    import lightgbm as lgb
    import optuna
    from sklearn.metrics import average_precision_score

    config = config or PropensityConfig(random_seed=random_seed)
    encoded, encoded_columns, feature_audit = _encode_lgbm_frames([train_frame, validation_frame, score_frame], feature_columns, config)
    train_matrix, validation_matrix, score_matrix = encoded
    y_train = train_frame["target"].astype(int).to_numpy()
    y_validation = validation_frame["target"].astype(int).to_numpy()
    positive_count = max(int(y_train.sum()), 1)
    negative_count = max(len(y_train) - int(y_train.sum()), 1)
    scale_pos_weight = negative_count / positive_count

    def pruning_callback(trial: optuna.Trial):
        def callback(environment: Any) -> None:
            for data_name, metric_name, value, _ in environment.evaluation_result_list:
                if data_name == "valid_0" and metric_name in {"aucpr", "average_precision"}:
                    trial.report(float(value), environment.iteration)
                    if trial.should_prune():
                        raise optuna.TrialPruned()
        callback.order = 10
        callback.before_iteration = False
        return callback

    def objective(trial: optuna.Trial) -> float:
        parameters = {
            "objective": "binary",
            "verbosity": -1,
            "n_estimators": trial.suggest_int("n_estimators", 150, 1000),
            "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.15, log=True),
            "num_leaves": trial.suggest_int("num_leaves", 8, 128, log=True),
            "max_depth": trial.suggest_int("max_depth", 3, 12),
            "min_child_samples": trial.suggest_int("min_child_samples", 20, 300),
            "subsample": trial.suggest_float("subsample", 0.65, 1.0),
            "colsample_bytree": trial.suggest_float("colsample_bytree", 0.65, 1.0),
            "reg_alpha": trial.suggest_float("reg_alpha", 1e-8, 10.0, log=True),
            "reg_lambda": trial.suggest_float("reg_lambda", 1e-8, 10.0, log=True),
            "scale_pos_weight": scale_pos_weight,
            "random_state": random_seed,
            "n_jobs": -1,
        }
        model = lgb.LGBMClassifier(**parameters)
        model.fit(
            train_matrix,
            y_train,
            eval_set=[(validation_matrix, y_validation)],
            eval_names=["valid_0"],
            eval_metric="aucpr",
            callbacks=[lgb.early_stopping(80, verbose=False), pruning_callback(trial)],
        )
        best_iteration = int(model.best_iteration_ or parameters["n_estimators"])
        trial.set_user_attr("best_iteration", best_iteration)
        train_probability = model.predict_proba(train_matrix, num_iteration=best_iteration)[:, 1]
        test_probability = model.predict_proba(validation_matrix, num_iteration=best_iteration)[:, 1]
        oot_probability = model.predict_proba(score_matrix, num_iteration=best_iteration)[:, 1]
        train_metrics = evaluate_predictions(train_frame["target"], train_probability, config.top_k)
        test_metrics = evaluate_predictions(validation_frame["target"], test_probability, config.top_k)
        oot_metrics = evaluate_predictions(score_frame["target"], oot_probability, config.top_k)
        trial.set_user_attr("train_pr_auc", train_metrics["pr_auc"])
        trial.set_user_attr("train_roc_auc", train_metrics["roc_auc"])
        trial.set_user_attr("train_brier", train_metrics["brier"])
        trial.set_user_attr("test_pr_auc", test_metrics["pr_auc"])
        trial.set_user_attr("test_roc_auc", test_metrics["roc_auc"])
        trial.set_user_attr("test_brier", test_metrics["brier"])
        trial.set_user_attr("oot_pr_auc", oot_metrics["pr_auc"])
        trial.set_user_attr("oot_roc_auc", oot_metrics["roc_auc"])
        trial.set_user_attr("oot_brier", oot_metrics["brier"])
        return float(test_metrics["pr_auc"])

    sampler = optuna.samplers.TPESampler(seed=random_seed)
    pruner = optuna.pruners.MedianPruner(n_startup_trials=5, n_warmup_steps=50, interval_steps=10)
    study = optuna.create_study(direction="maximize", sampler=sampler, pruner=pruner)
    study.optimize(objective, n_trials=n_trials, timeout=timeout, show_progress_bar=False)
    if not study.best_trials:
        raise RuntimeError("Optuna gecerli bir complete trial uretemedi; tum trial'lar prune veya fail oldu")

    best_parameters = dict(study.best_trial.params)
    best_parameters.update({
        "objective": "binary",
        "verbosity": -1,
        "scale_pos_weight": scale_pos_weight,
        "random_state": random_seed,
        "n_jobs": -1,
        "n_estimators": int(study.best_trial.user_attrs.get("best_iteration", best_parameters["n_estimators"])),
    })
    tuned_model = lgb.LGBMClassifier(**best_parameters)
    tuned_model.fit(
        train_matrix,
        y_train,
        eval_set=[(validation_matrix, y_validation)],
        eval_names=["valid_0"],
        eval_metric="aucpr",
        callbacks=[lgb.early_stopping(80, verbose=False)],
    )
    tuned_iteration = int(tuned_model.best_iteration_ or best_parameters["n_estimators"])
    tuned_metrics: list[dict[str, Any]] = []
    for split_name, split_frame, split_matrix in (("train", train_frame, train_matrix), ("test", validation_frame, validation_matrix), ("oot", score_frame, score_matrix)):
        probability = tuned_model.predict_proba(split_matrix, num_iteration=tuned_iteration)[:, 1]
        tuned_metrics.append({"split": split_name, "evaluation_stage": "best_trial_model", "fit_scope": "train_only", **evaluate_predictions(split_frame["target"], probability, config.top_k)})
    fit_frame = pd.concat([train_frame, validation_frame], ignore_index=True)
    fit_matrix = pd.concat([train_matrix, validation_matrix], ignore_index=True)
    final_model = lgb.LGBMClassifier(**best_parameters)
    final_model.fit(fit_matrix, fit_frame["target"].astype(int).to_numpy(), callbacks=[lgb.log_evaluation(0)])
    score_probability = final_model.predict_proba(score_matrix)[:, 1]
    metrics = evaluate_predictions(score_frame["target"], score_probability)
    predictions = score_frame[["musteri_id", "anchor_month", "product_class"]].copy()
    predictions["probability"] = score_probability
    final_metrics: list[dict[str, Any]] = tuned_metrics
    for split_name, split_frame, split_matrix in (("train", train_frame, train_matrix), ("test", validation_frame, validation_matrix), ("oot", score_frame, score_matrix)):
        probability = final_model.predict_proba(split_matrix)[:, 1]
        final_metrics.append({"split": split_name, "evaluation_stage": "final_selected_model", "fit_scope": "train_plus_test", **evaluate_predictions(split_frame["target"], probability, config.top_k)})
    metrics_frame = pd.DataFrame(final_metrics)
    trial_rows: list[dict[str, Any]] = []
    for trial in study.trials:
        trial_rows.append({
            "trial_number": trial.number,
            "trial_value_test_pr_auc": trial.value,
            "trial_state": str(trial.state).split(".")[-1],
            "trial_datetime_start": trial.datetime_start,
            "trial_datetime_complete": trial.datetime_complete,
            "trial_best_iteration": trial.user_attrs.get("best_iteration"),
            **{key: trial.user_attrs.get(key, np.nan) for key in ("train_pr_auc", "train_roc_auc", "train_brier", "test_pr_auc", "test_roc_auc", "test_brier", "oot_pr_auc", "oot_roc_auc", "oot_brier")},
            **{f"param_{key}": value for key, value in trial.params.items()},
        })
    trial_frame = pd.DataFrame(trial_rows)
    return {"model": final_model, "encoded_columns": encoded_columns, "best_parameters": best_parameters, "study": study, "feature_audit": feature_audit}, metrics_frame, predictions, trial_frame


def run_optuna_grid(
    targets: pd.DataFrame,
    features: pd.DataFrame,
    config: PropensityConfig,
    n_trials: int = 30,
    timeout_per_configuration: int | None = 600,
    max_configurations: int | None = None,
    feature_columns: list[str] | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, Any], pd.DataFrame]:
    """Run independent pruned Optuna-LightGBM models for every segment/configuration."""

    feature_columns = feature_columns or [column for column in features.columns if column not in KEY_COLUMNS and column != config.segment_column]
    missing_features = sorted(set(feature_columns) - set(features.columns))
    if missing_features:
        raise ValueError(f"Feature columns features tablosunda yok: {missing_features}")
    forbidden_features = sorted(set(feature_columns) & (set(KEY_COLUMNS) | {config.segment_column}))
    if forbidden_features:
        raise ValueError(f"Key veya segment kolonlari model feature olamaz: {forbidden_features}")
    if config.segment_column not in features.columns:
        raise ValueError(f"Segment kolonu bulunamadi: {config.segment_column}")
    audit = target_audit(targets)
    if max_configurations is not None:
        audit = audit.head(max_configurations)
    available_segments = features[config.segment_column].dropna().map(_normalize_segment_value).unique().tolist()
    requested_segments = config.segment_values if config.segment_values is not None else tuple(sorted(available_segments, key=str))
    segment_values = [_normalize_segment_value(value) for value in requested_segments]
    metric_rows: list[dict[str, Any]] = []
    prediction_rows: list[pd.DataFrame] = []
    trial_rows: list[pd.DataFrame] = []
    model_registry: dict[str, Any] = {}
    for row in audit.itertuples(index=False):
        threshold = None if pd.isna(row.r_threshold) else float(row.r_threshold)
        mask = _configuration_mask(targets, row.model_type, row.product_class, int(row.x_window), int(row.y_window), threshold)
        base_selected = targets.loc[mask & targets["eligible"]].merge(features, on=list(KEY_COLUMNS), how="left", validate="one_to_one").dropna(subset=["target"])
        for segment in segment_values:
            model_key = f"segment_{segment}_{row.model_type}_{row.product_class}_x{row.x_window}_y{row.y_window}_r{threshold}"
            selected = base_selected[base_selected[config.segment_column].map(_normalize_segment_value).eq(segment)].copy()
            common = {"model_key": model_key, "segment": segment, "model_name": "optuna_lightgbm", "model_type": row.model_type, "product_class": row.product_class, "x_window": row.x_window, "y_window": row.y_window, "r_threshold": threshold}
            eligible_count = len(selected)
            positive_count = int(selected["target"].sum()) if not selected.empty else 0
            if eligible_count < config.min_eligible or positive_count < config.min_positive:
                metric_rows.append({**common, "split": "skipped", "evaluation_stage": "configuration_gate", "fit_scope": "not_fitted", "error": f"segment veri esigi saglanmadi: eligible={eligible_count}, positive={positive_count}"})
                continue
            try:
                split = make_time_split(
                    selected,
                    test_months=config.test_months,
                    oot_months=config.oot_months,
                    min_positive_train=config.min_positive_train,
                    min_positive_test=config.min_positive_test,
                    min_positive_oot=config.min_positive_oot,
                )
            except ValueError as error:
                metric_rows.append({**common, "split": "skipped", "evaluation_stage": "time_split", "fit_scope": "not_fitted", "error": str(error)})
                continue
            train = selected[selected["anchor_month"].isin(split["train_months"])].copy()
            validation = selected[selected["anchor_month"].isin(split["test_months"])].copy()
            oot = selected[selected["anchor_month"].isin(split["oot_months"])].copy()
            if min(frame["target"].nunique() for frame in (train, validation, oot)) < 2:
                metric_rows.append({**common, "split": "skipped", "evaluation_stage": "class_balance_gate", "fit_scope": "not_fitted", "error": "train, test veya OOT splitinde iki target sinifi bulunmuyor"})
                continue
            try:
                bundle, split_metrics, predictions, trials = fit_optuna_lightgbm(train, validation, oot, feature_columns, n_trials=n_trials, timeout=timeout_per_configuration, random_seed=config.random_seed, config=config)
            except (ImportError, RuntimeError, ValueError) as error:
                metric_rows.append({**common, "split": "error", "evaluation_stage": "fit", "fit_scope": "not_fitted", "error": str(error)})
                continue
            period_info = {"train_period_start": train["anchor_month"].min(), "train_period_end": train["anchor_month"].max(), "test_period": validation["anchor_month"].min(), "oot_period": oot["anchor_month"].min()}
            for metric_row in split_metrics.to_dict("records"):
                metric_rows.append({**common, **period_info, **metric_row})
            predictions = predictions.assign(model_key=model_key, model_name="optuna_lightgbm", segment=segment, model_type=row.model_type, x_window=row.x_window, y_window=row.y_window, r_threshold=threshold)
            prediction_rows.append(predictions)
            trials = trials.assign(model_key=model_key, segment=segment, model_type=row.model_type, product_class=row.product_class, x_window=row.x_window, y_window=row.y_window, r_threshold=threshold, **period_info)
            trial_rows.append(trials)
            bundle["feature_audit"] = bundle["feature_audit"].assign(model_key=model_key, segment=segment, model_type=row.model_type, product_class=row.product_class, x_window=row.x_window, y_window=row.y_window, r_threshold=threshold, **period_info)
            model_registry[model_key] = bundle
    metrics_frame = pd.DataFrame(metric_rows)
    predictions_frame = pd.concat(prediction_rows, ignore_index=True) if prediction_rows else pd.DataFrame()
    trials_frame = pd.concat(trial_rows, ignore_index=True) if trial_rows else pd.DataFrame()
    return metrics_frame, predictions_frame, model_registry, trials_frame


def rank_campaign(scored_frame: pd.DataFrame, probability_column: str = "probability", capacity: int | None = None) -> pd.DataFrame:
    """Rank customers and apply an optional campaign capacity."""

    result = scored_frame.sort_values(probability_column, ascending=False).reset_index(drop=True).copy()
    result["rank"] = np.arange(1, len(result) + 1)
    result["decile"] = pd.qcut(result["rank"], q=min(10, len(result)), labels=False, duplicates="drop") + 1
    result["selected"] = True if capacity is None else result["rank"] <= capacity
    return result


def make_synthetic_fixture(seed: int = 42, customer_count: int = 120) -> SourceBundle:
    """Small deterministic fixture for notebook smoke tests, not training data."""

    rng = np.random.default_rng(seed)
    months = pd.date_range("2025-06-01", "2026-06-01", freq="MS")
    customers = pd.DataFrame({"musteri_id": [f"M{index:05d}" for index in range(customer_count)], "segment": (np.arange(customer_count) % 5) + 1})
    activity_rows: list[dict[str, Any]] = []
    flow_rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []
    for customer_index, customer_id in enumerate(customers["musteri_id"]):
        for month_index, month in enumerate(months):
            monthly_activity = {}
            for product_index, product_class in enumerate(PRODUCT_CLASSES):
                active = int((customer_index + month_index + product_index) % 5 != 0)
                if product_class == "para_piyasasi" and customer_index % 4 == 0 and month_index >= 9:
                    active = 0
                elif product_class == "nitelikli" and customer_index % 4 == 3 and month_index >= 9:
                    active = 0
                elif product_class == "para_piyasasi" and customer_index % 7 == 0 and month_index >= 7:
                    active = int(month_index >= 10)
                monthly_activity[ACTIVITY_WIDE_COLUMNS[product_class]] = active
                has_future_newsell_signal = month_index >= 9 and (customer_index + product_index) % 4 != 0
                buy = float(rng.integers(0, 9000) if active or has_future_newsell_signal else 0)
                sell = float(rng.integers(0, 2500) if active else 0)
                fund_value = float(5000 + customer_index * 100 + month_index * 1500 + product_index * 2500) if active else 0.0
                flow_rows.append({"musteri_id": customer_id, "month": month, "product_class": product_class, "buy_amount": buy, "sell_amount": sell, "fund_value": fund_value})
            activity_rows.append({"musteri_id": customer_id, "month": month, **monthly_activity})
            feature_rows.append({"musteri_id": customer_id, "month": month, "monthly_income": float(25000 + customer_index * 100)})
    inflation = pd.DataFrame({"month": months, "inflation_rate": np.full(len(months), 0.025)})
    return SourceBundle(customers, pd.DataFrame(activity_rows), pd.DataFrame(flow_rows), pd.DataFrame(feature_rows), inflation)


def make_rare_event_fixture(seed: int = 42, customer_count: int = 10000, target_positive_rate: float = 0.001) -> SourceBundle:
    """Create a larger five-segment fixture with approximately rare-event targets."""

    if customer_count < 5 or customer_count % 5:
        raise ValueError("customer_count bes segmente esit dagilabilmeli")
    if not 0 < target_positive_rate < 1:
        raise ValueError("target_positive_rate 0 ile 1 arasinda olmali")
    rng = np.random.default_rng(seed)
    months = pd.date_range("2025-01-01", "2026-06-01", freq="MS")
    customer_ids = [f"R{index:06d}" for index in range(customer_count)]
    customers = pd.DataFrame({"musteri_id": customer_ids, "segment": (np.arange(customer_count) % 5) + 1})
    event_buy: dict[tuple[int, int, int], float] = {}
    segment_size = customer_count // 5
    for segment_index in range(5):
        segment_start = segment_index * segment_size
        segment_indexes = np.arange(segment_start, segment_start + segment_size)
        for product_index in range(len(PRODUCT_CLASSES)):
            newsell_pool = segment_indexes[(segment_indexes + product_index) % 2 == 0]
            upsell_pool = segment_indexes[(segment_indexes + product_index) % 2 == 1]
            event_count_new = max(1, int(round(len(newsell_pool) * target_positive_rate)))
            event_count_up = max(1, int(round(len(upsell_pool) * target_positive_rate)))
            for anchor_index in range(3, len(months) - 1):
                future_index = anchor_index + 1
                for event_index, pool, count in ((anchor_index, newsell_pool, event_count_new), (anchor_index + len(months), upsell_pool, event_count_up)):
                    source_anchor = event_index if event_index < len(months) else event_index - len(months)
                    chosen_pool = pool[(source_anchor * 37 + segment_index * 11 + product_index * 5) % len(pool)]
                    event_buy[(int(chosen_pool), product_index, future_index)] = 2500.0

    activity_rows: list[dict[str, Any]] = []
    flow_rows: list[dict[str, Any]] = []
    feature_rows: list[dict[str, Any]] = []
    for customer_index, customer_id in enumerate(customer_ids):
        segment_index = customer_index // segment_size
        for month_index, month in enumerate(months):
            monthly_activity: dict[str, int] = {}
            for product_index, product_class in enumerate(PRODUCT_CLASSES):
                newsell_pool_member = (customer_index + product_index) % 2 == 0
                active = int(not newsell_pool_member or month_index < 3)
                monthly_activity[ACTIVITY_WIDE_COLUMNS[product_class]] = active
                base_buy = float(rng.integers(0, 8)) if active else 0.0
                buy = event_buy.get((customer_index, product_index, month_index), base_buy)
                sell = float(rng.integers(0, 3)) if active else 0.0
                fund_value = float(10000 + segment_index * 500 + customer_index % 100 * 25 + month_index * 100) if active else 0.0
                flow_rows.append({"musteri_id": customer_id, "month": month, "product_class": product_class, "buy_amount": buy, "sell_amount": sell, "fund_value": fund_value})
            activity_rows.append({"musteri_id": customer_id, "month": month, **monthly_activity})
            feature_rows.append({"musteri_id": customer_id, "month": month, "monthly_income": float(25000 + segment_index * 5000 + customer_index % 200 * 100)})
    inflation = pd.DataFrame({"month": months, "inflation_rate": np.full(len(months), 0.02)})
    return SourceBundle(customers, pd.DataFrame(activity_rows), pd.DataFrame(flow_rows), pd.DataFrame(feature_rows), inflation)


def _read_source_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".parquet":
        return pd.read_parquet(path)
    if path.suffix.lower() in {".csv", ".txt"}:
        return pd.read_csv(path)
    raise ValueError(f"Desteklenmeyen source formatı: {path.suffix}")


def _read_configured_table(root: Path, filename: str, name: str) -> pd.DataFrame:
    path = root / filename
    if not path.exists():
        raise FileNotFoundError(f"{name} tablosu bulunamadı: {path}")
    return _read_source_file(path)


def _load_manual_tables(root: Path, config: PropensityConfig) -> SourceBundle:
    """Load the user-owned monthly input, activity, fund and transaction tables."""

    required_files = {
        "input_table_file": config.input_table_file,
        "activity_table_file": config.activity_table_file,
        "fund_table_file": config.fund_table_file,
        "transaction_table_file": config.transaction_table_file,
        "inflation_table_file": config.inflation_table_file,
    }
    missing_files = [name for name, value in required_files.items() if not value]
    if missing_files:
        raise ValueError(f"Manuel tablo dosyaları config'te eksik: {missing_files}")

    input_source = _read_configured_table(root, config.input_table_file, "input")
    input_customer = config.input_table_customer_column
    input_date = config.input_table_date_column
    _require_columns(input_source, [input_customer, input_date], "input")
    input_source = input_source.rename(columns={input_customer: "musteri_id", input_date: "month"})
    input_source["musteri_id"] = input_source["musteri_id"].astype(str)
    input_source["month"] = _month_start(input_source["month"])
    if input_source["month"].isna().any():
        raise ValueError("input tarih kolonunda parse edilemeyen değer var")
    customers = input_source[["musteri_id"] + (["segment"] if "segment" in input_source.columns else [])].drop_duplicates("musteri_id")
    if "segment" not in customers:
        customers["segment"] = "unknown"

    activity_source = _read_configured_table(root, config.activity_table_file, "aktivite")
    activity_customer = config.activity_table_customer_column
    activity_date = config.activity_table_date_column
    _require_columns(activity_source, [activity_customer, activity_date, config.activity_ppf_flag_column, config.activity_nf_flag_column], "aktivite")
    activity = activity_source.rename(columns={activity_customer: "musteri_id", activity_date: "month", config.activity_ppf_flag_column: "ppf_aktif", config.activity_nf_flag_column: "nf_aktif"})
    activity = activity[["musteri_id", "month", "ppf_aktif", "nf_aktif"]].copy()
    activity["musteri_id"] = activity["musteri_id"].astype(str)
    activity["month"] = _month_start(activity["month"])

    fund_source = _read_configured_table(root, config.fund_table_file, "fon tutar")
    fund_customer = config.fund_table_customer_column
    fund_date = config.fund_table_date_column
    _require_columns(fund_source, [fund_customer, fund_date, config.fund_table_product_flag_column, config.fund_table_value_column], "fon tutar")
    fund = fund_source.rename(columns={fund_customer: "musteri_id", fund_date: "month", config.fund_table_product_flag_column: "para_flg", config.fund_table_value_column: "fund_value"})
    fund = fund[["musteri_id", "month", "para_flg", "fund_value"]].copy()
    fund["musteri_id"] = fund["musteri_id"].astype(str)
    fund["month"] = _month_start(fund["month"])
    fund["product_class"] = np.where(pd.to_numeric(fund["para_flg"], errors="coerce").eq(1), "para_piyasasi", "nitelikli")
    fund["fund_value"] = pd.to_numeric(fund["fund_value"], errors="coerce").fillna(0.0)
    fund = fund.groupby(["musteri_id", "month", "product_class"], as_index=False)["fund_value"].sum()

    transaction_source = _read_configured_table(root, config.transaction_table_file, "alım satım")
    transaction_customer = config.transaction_table_customer_column
    transaction_date = config.transaction_table_date_column
    _require_columns(transaction_source, [transaction_customer, transaction_date, config.transaction_table_product_flag_column, config.transaction_table_buy_column, config.transaction_table_sell_column], "alım satım")
    transactions = transaction_source.rename(columns={transaction_customer: "musteri_id", transaction_date: "month", config.transaction_table_product_flag_column: "ppf_flg", config.transaction_table_buy_column: "buy_amount", config.transaction_table_sell_column: "sell_amount"})
    transactions = transactions[["musteri_id", "month", "ppf_flg", "buy_amount", "sell_amount"]].copy()
    transactions["musteri_id"] = transactions["musteri_id"].astype(str)
    transactions["month"] = _month_start(transactions["month"])
    transactions["product_class"] = np.where(pd.to_numeric(transactions["ppf_flg"], errors="coerce").eq(1), "para_piyasasi", "nitelikli")
    transactions["buy_amount"] = pd.to_numeric(transactions["buy_amount"], errors="coerce").fillna(0.0)
    transactions["sell_amount"] = pd.to_numeric(transactions["sell_amount"], errors="coerce").fillna(0.0)
    transactions = transactions.groupby(["musteri_id", "month", "product_class"], as_index=False)[["buy_amount", "sell_amount"]].sum()
    flows = fund.merge(transactions, on=["musteri_id", "month", "product_class"], how="outer", validate="one_to_one")
    for column in ("fund_value", "buy_amount", "sell_amount"):
        flows[column] = pd.to_numeric(flows[column], errors="coerce").fillna(0.0)

    reserved = {input_customer, input_date, "musteri_id", "month", "segment"}
    feature_columns = list(config.input_table_feature_columns) if config.input_table_feature_columns else [column for column in input_source.columns if column not in reserved]
    feature_columns = [column for column in feature_columns if column in input_source.columns]
    monthly_features = input_source[["musteri_id", "month", *feature_columns]].copy()
    monthly_features = monthly_features.groupby(["musteri_id", "month"], as_index=False)[feature_columns].first() if feature_columns else pd.DataFrame(columns=["musteri_id", "month"])
    inflation = pd.DataFrame()
    if config.inflation_table_file:
        inflation_source = _read_configured_table(root, config.inflation_table_file, "enflasyon")
        _require_columns(inflation_source, [config.inflation_date_column, config.inflation_value_column], "enflasyon")
        inflation = inflation_source.rename(columns={config.inflation_date_column: "month", config.inflation_value_column: "inflation_rate"})[["month", "inflation_rate"]]
    return adapt_source_bundle(SourceBundle(customers, activity, flows, monthly_features, inflation))


def _load_wide_input_table(root: Path, config: PropensityConfig) -> SourceBundle:
    """Convert one wide customer-month table into leakage-safe source tables."""

    table_path = root / config.input_table_file if config.input_table_file else None
    if table_path is None or not table_path.exists():
        raise FileNotFoundError(f"Geniş input tablosu bulunamadı: {table_path}")
    source = _read_source_file(table_path)
    customer = config.input_table_customer_column
    date = config.input_table_date_column
    _require_columns(source, [customer, date], "input_table")
    source = source.rename(columns={customer: "musteri_id", date: "month"})
    source["musteri_id"] = source["musteri_id"].astype(str)
    source["month"] = _month_start(source["month"])
    if source["month"].isna().any():
        raise ValueError("input_table tarih kolonunda parse edilemeyen değer var")
    if source.duplicated(["musteri_id", "month"] + ([config.input_table_product_column] if config.input_table_product_column and config.input_table_product_column in source else [])).any():
        raise ValueError("input_table anahtarında duplicate var; aylık müşteri/ürün satırı tekil olmalı")

    segment = source[["musteri_id"]].copy()
    if "segment" in source.columns:
        segment["segment"] = source["segment"].astype(str)
    else:
        segment["segment"] = "unknown"
    customers = segment.drop_duplicates("musteri_id")
    product_column = config.input_table_product_column
    activity_mapping = config.input_table_activity_columns
    activity_parts: list[pd.DataFrame] = []
    if product_column and product_column in source.columns:
        product_values = source[product_column].astype(str)
        for product_class, activity_column in activity_mapping.items():
            if activity_column in source.columns:
                part = source[["musteri_id", "month", activity_column]].drop_duplicates(["musteri_id", "month"], keep="first").rename(columns={activity_column: "active_flag"})
                part["product_class"] = product_class
                activity_parts.append(part)
    else:
        for product_class, activity_column in activity_mapping.items():
            _require_columns(source, [activity_column], "input_table activity")
            part = source[["musteri_id", "month", activity_column]].drop_duplicates(["musteri_id", "month"], keep="first").rename(columns={activity_column: "active_flag"})
            part["product_class"] = product_class
            activity_parts.append(part)
    activity = pd.concat(activity_parts, ignore_index=True)
    flow_columns = [config.input_table_buy_column, config.input_table_sell_column, config.input_table_fund_value_column]
    _require_columns(source, flow_columns, "input_table flows")
    if product_column and product_column in source.columns:
        flows = source[["musteri_id", "month", product_column, *flow_columns]].rename(columns={product_column: "product_class", config.input_table_buy_column: "buy_amount", config.input_table_sell_column: "sell_amount", config.input_table_fund_value_column: "fund_value"})
    else:
        flow_parts = []
        for product_class in config.product_classes:
            part = source[["musteri_id", "month", *flow_columns]].copy()
            part["product_class"] = product_class
            flow_parts.append(part.rename(columns={config.input_table_buy_column: "buy_amount", config.input_table_sell_column: "sell_amount", config.input_table_fund_value_column: "fund_value"}))
        flows = pd.concat(flow_parts, ignore_index=True)
    reserved = {"musteri_id", "month", "segment", product_column, *activity_mapping.values(), *flow_columns}
    feature_columns = list(config.input_table_feature_columns) if config.input_table_feature_columns else [column for column in source.columns if column not in reserved]
    feature_columns = [column for column in feature_columns if column in source.columns]
    monthly_features = source[["musteri_id", "month", *feature_columns]].copy()
    if feature_columns:
        aggregation = config.input_table_feature_aggregation
        if aggregation not in {"first", "mean", "sum", "max", "min"}:
            raise ValueError("input_table_feature_aggregation first/mean/sum/max/min olmalı")
        if aggregation != "first":
            monthly_features = monthly_features.groupby(["musteri_id", "month"], as_index=False)[feature_columns].agg(aggregation)
        else:
            monthly_features = monthly_features.groupby(["musteri_id", "month"], as_index=False)[feature_columns].first()
    else:
        monthly_features = pd.DataFrame(columns=["musteri_id", "month"])
    inflation = pd.DataFrame()
    if config.inflation_table_file:
        inflation_path = root / config.inflation_table_file
        inflation = _read_source_file(inflation_path).rename(columns={config.inflation_date_column: "month", config.inflation_value_column: "inflation_rate"})
    return adapt_source_bundle(SourceBundle(customers, activity, flows, monthly_features, inflation))


def load_bundle_from_directory(data_root: str | Path, config: PropensityConfig | None = None) -> SourceBundle:
    """Load either a configured wide table or normalized CSV/Parquet sources."""

    root = Path(data_root)
    if config is not None and config.activity_table_file:
        return _load_manual_tables(root, config)
    if config is not None and config.input_table_file:
        return _load_wide_input_table(root, config)

    def read(name: str, optional: bool = False) -> pd.DataFrame:
        parquet_path = root / f"{name}.parquet"
        csv_path = root / f"{name}.csv"
        if parquet_path.exists():
            return pd.read_parquet(parquet_path)
        if csv_path.exists():
            return pd.read_csv(csv_path)
        if optional:
            return pd.DataFrame()
        raise FileNotFoundError(f"Beklenen kaynak bulunamadi: {parquet_path} veya {csv_path}")

    return adapt_source_bundle(SourceBundle(read("customers"), read("activity"), read("flows"), read("monthly_features", True), read("inflation", True)))